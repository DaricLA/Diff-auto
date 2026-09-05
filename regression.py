# -*- coding: utf-8 -*-
"""
L2 全场景回归测试 v3（单文件）：金标准 14 案例 + B1/B2/B3/B4/B5 全场景验证
双击 regression.exe -> out/RULES_VERDICT.txt + out/L2_REPORT.txt
（用例数据已内联，无外部模块依赖，打包即用）
"""
import os, sys, time, threading, traceback, re, zipfile, io, tempfile
from collections import Counter

BASE = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
if not getattr(sys, 'frozen', False):
    sys.path.insert(0, BASE)

import main
from main import OpenpyxlComparer, DEFAULT_CHECK_OPTIONS, CheckProject, CheckRule, CheckItemConfig

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
try:
    from openpyxl.formatting.rule import CellIsRule
except Exception:
    CellIsRule = None
try:
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    HAS_RICH = True
except Exception:
    HAS_RICH = False
try:
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage
    HAS_PIL = True
except Exception:
    HAS_PIL = False

CHECK_TYPES = ['value', 'formula', 'rich_text', 'font', 'fill', 'border',
               'alignment', 'number_format', 'merged_cells', 'row_height',
               'col_width', 'images', 'conditional_format']


# ============================================================
# 一、金标准回归（原 14 案例）
# ============================================================
def _save_pair(name, old_wb, new_wb, tmpdir):
    o = os.path.join(tmpdir, name + '_old.xlsx')
    n = os.path.join(tmpdir, name + '_new.xlsx')
    old_wb.save(o)
    new_wb.save(n)
    return o, n

def _base():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws['A1'] = 1; ws['B1'] = 'x'; ws['A2'] = 10
    return wb, ws

def c_value_changed():
    old, ws = _base(); new, ws2 = _base(); ws2['A1'] = 2
    return old, new, {'must': [('内容变化', 'A1')], 'absent': []}, {}

def c_identical():
    old, _ = _base(); new, _ = _base()
    return old, new, {'must': [], 'absent': ['内容变化', '公式变化', '字体变化', '填充变化',
                                              '边框变化', '对齐变化', '数字格式变化']}, {}

def c_numfmt_changed():
    old, ws = _base(); new, ws2 = _base()
    ws['A1'].number_format = '0'; ws2['A1'].number_format = '0.00'
    return old, new, {'must': [('数字格式变化', 'A1')], 'absent': ['内容变化']}, {}

def c_numfmt_gate_off():
    old, ws = _base(); new, ws2 = _base()
    ws['A1'].number_format = '0'; ws2['A1'].number_format = '0.00'
    return old, new, {'must': [], 'absent': ['数字格式变化']}, {'number_format': False}

def c_font_name_changed():
    old, ws = _base(); new, ws2 = _base()
    ws['A1'].font = Font(name='微软雅黑'); ws2['A1'].font = Font(name='宋体')
    return old, new, {'must': [('字体变化', 'A1')], 'absent': []}, {}

def c_font_size_changed():
    old, ws = _base(); new, ws2 = _base()
    ws['A1'].font = Font(size=11); ws2['A1'].font = Font(size=14)
    return old, new, {'must': [('字体变化', 'A1')], 'absent': []}, {}

def c_fill_changed():
    old, ws = _base(); new, ws2 = _base()
    ws['A1'].fill = PatternFill(fill_type='solid', start_color='FFFF00')
    ws2['A1'].fill = PatternFill(fill_type='solid', start_color='0000FF')
    return old, new, {'must': [('填充变化', 'A1')], 'absent': []}, {}

def c_border_changed():
    old, ws = _base(); new, ws2 = _base()
    thin = Side(style='thin', color='000000')
    ws2['A1'].border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return old, new, {'must': [('边框变化', 'A1')], 'absent': []}, {}

def c_align_changed():
    old, ws = _base(); new, ws2 = _base()
    ws2['A1'].alignment = Alignment(horizontal='center')
    return old, new, {'must': [('对齐变化', 'A1')], 'absent': []}, {}

def c_rowheight_changed():
    old, ws = _base(); new, ws2 = _base()
    ws2.row_dimensions[2].height = 30
    return old, new, {'must': [('行高变化', 'A2')], 'absent': []}, {}

def c_colwidth_changed():
    old, ws = _base(); new, ws2 = _base()
    ws2.column_dimensions['A'].width = 20
    return old, new, {'must': [('列宽变化', 'A1')], 'absent': []}, {}

def c_merged_changed():
    old, ws = _base(); new, ws2 = _base()
    ws.merge_cells('A1:B1')
    return old, new, {'must': [('合并删除', 'A1')], 'absent': []}, {}

def c_formula_changed():
    old, ws = _base(); new, ws2 = _base()
    ws['A2'] = '=1+1'; ws2['A2'] = '=2+2'
    return old, new, {'must': [('公式变化', 'A2')], 'absent': []}, {}

def c_sheet_added():
    old, ws = _base(); new, ws2 = _base()
    new.create_sheet('新增表')
    return old, new, {'must': [], 'absent': []}, {}

CASES = [
    ('v01 值变化检测', c_value_changed),
    ('v02 完全相同应零差异', c_identical),
    ('v03 数字格式变化检测', c_numfmt_changed),
    ('v04 数字格式检测关闭(豁免)', c_numfmt_gate_off),
    ('v05 字体名变化检测', c_font_name_changed),
    ('v06 字号变化检测', c_font_size_changed),
    ('v07 填充色变化检测', c_fill_changed),
    ('v08 边框变化检测', c_border_changed),
    ('v09 对齐变化检测', c_align_changed),
    ('v10 行高变化检测', c_rowheight_changed),
    ('v11 列宽变化检测', c_colwidth_changed),
    ('v12 合并单元格删除检测', c_merged_changed),
    ('v13 公式变化检测', c_formula_changed),
    ('v14 新增sheet检测', c_sheet_added),
]

def run_one(name, builder, tmpdir):
    try:
        old, new, expect, opts = builder()
        o, n = _save_pair(name.replace(' ', '_'), old, new, tmpdir)
        copt = dict(DEFAULT_CHECK_OPTIONS); copt.update(opts or {})
        cmp = OpenpyxlComparer(o, n, log_callback=lambda m: None,
                               progress_callback=lambda v, s=None: None,
                               progress_mode_fn=lambda m: None,
                               check_options=copt, stop_event=threading.Event())
        ok = cmp.run()
        if not ok:
            return {'ok': False, 'detail': '引擎返回失败', 'counts': {}}
        found = [(d.get('type'), str(d.get('address', '')).split(':')[0].upper()) for d in cmp.diffs]
        counts = Counter(t for t, a in found)
        problems = []
        for t, a in expect.get('must', []):
            if not any(ft == t and (fa == a.upper() or fa.startswith(a.upper())) for ft, fa in found):
                problems.append('缺少期望差异 %s@%s' % (t, a))
        for t in expect.get('absent', []):
            if any(ft == t for ft, fa in found):
                problems.append('不应报但报了 %s' % t)
        return {'ok': not problems, 'detail': '; '.join(problems) if problems else '符合预期',
                'counts': counts, 'found': found[:20]}
    except Exception as e:
        return {'ok': False, 'detail': '异常: %s | %s' % (e, traceback.format_exc()[:400]), 'counts': {}}

# ============================================================
# 二、L2 用例数据源/工具
# ============================================================
def _range_ds(sheet, anchor, r_off, c_off, r_cnt, c_cnt, name='L2'):
    return {'name': name, 'sheet': sheet, 'anchor': {'text': anchor}, 'search_in': 'all',
            'mode': 'range',
            'target': {'row_offset': r_off, 'col_offset': c_off,
                       'row_count': r_cnt, 'col_count': c_cnt}}


def _shift_ds(sheet, anchor, shift_offset, rows, name='L2'):
    return {'name': name, 'sheet': sheet, 'anchor': {'text': anchor}, 'search_in': 'all',
            'mode': 'shift',
            'header_target': {'row_offset': 0, 'col_offset': 0, 'row_count': 1, 'col_count': 4},
            'rows': rows, 'shift_offset': shift_offset}


def _fill_diff(ow, nw, r, c, kinds):
    """在普通布局 (r,c) 制造指定类型的差异（同地址）"""
    cell = '%s%d' % (get_column_letter(c), r)
    if 'value' in kinds:
        ow.cell(r, c, 10); nw.cell(r, c, 20)
    if 'formula' in kinds:
        ow.cell(r, c, '=1+1'); nw.cell(r, c, '=2+2')
    if 'rich_text' in kinds and HAS_RICH:
        ow.cell(r, c, 'AB')
        nw.cell(r, c, CellRichText(TextBlock(InlineFont(b=True), 'AB')))
    if 'font' in kinds:
        ow.cell(r, c, 'x').font = Font(name='宋体')
        nw.cell(r, c, 'x').font = Font(name='微软雅黑')
    if 'fill' in kinds:
        ow.cell(r, c).fill = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
        nw.cell(r, c).fill = PatternFill(fill_type='solid', start_color='0000FF', end_color='0000FF')
    if 'border' in kinds:
        thin = Side(style='thin', color='000000')
        nw.cell(r, c).border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if 'alignment' in kinds:
        nw.cell(r, c).alignment = Alignment(horizontal='center')
    if 'number_format' in kinds:
        ow.cell(r, c).number_format = '0'
        nw.cell(r, c).number_format = '0.00'
    if 'merged_cells' in kinds:
        nw.merge_cells('%s%d:%s%d' % (get_column_letter(c), r, get_column_letter(c), r + 1))
    if 'row_height' in kinds:
        nw.row_dimensions[r].height = 30
    if 'col_width' in kinds:
        nw.column_dimensions[get_column_letter(c)].width = 25
    if 'images' in kinds and HAS_PIL:
        _fpx = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
        _pngx = _fpx.name; _fpx.close()
        PILImage.new('RGB', (10, 10), (255, 0, 0)).save(_pngx, format='PNG')
        nw.add_image(XLImage(_pngx), cell)
    if 'conditional_format' in kinds and CellIsRule is not None:
        nw.conditional_formatting.add(cell, CellIsRule(operator='greaterThan', formula=['5'],
                                                       fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')))


def _set_theme_accent1(path, hex6):
    """替换 xlsx 主题表 accent1 颜色（zip 重写 theme1.xml）"""
    tmp = path + '.tmp'
    zin = zipfile.ZipFile(path, 'r')
    zout = zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED)
    for item in zin.infolist():
        data = zin.read(item.filename)
        if item.filename == 'xl/theme/theme1.xml':
            try:
                s = data.decode('utf-8')
                s2 = re.sub(r'(<a:accent1>.*?<a:srgbClr val=")[0-9A-Fa-f]{6}(?="/>)',
                            r'\g<1>' + hex6, s, flags=re.S)
                data = s2.encode('utf-8')
            except Exception:
                pass
        zout.writestr(item, data)
    zin.close()
    zout.close()
    os.replace(tmp, path)


# ============================================================
# 三、L2 通用工具（断言/COM/引擎执行）
# ============================================================
def l2_make_project(rules):
    return CheckProject('L2', '1.0', [CheckRule(rule_name='L2', data_source=ds,
                                                checks=[CheckItemConfig(check_type=t, enabled=True, expect=e)
                                                        for (t, e) in (checks or [])])
                                       for ds, checks in rules])


def l2_save(data, name, tmpdir):
    o = os.path.join(tmpdir, 'l2_' + re.sub(r'[\\/:*?"<>|]', '_', name) + '_old.xlsx')
    n = os.path.join(tmpdir, 'l2_' + re.sub(r'[\\/:*?"<>|]', '_', name) + '_new.xlsx')
    data['old'].save(o)
    data['new'].save(n)
    return o, n


def diff_at(diffs, sheet, addr, types=None):
    a = str(addr).upper()
    for d in diffs:
        if d.get('sheet') != sheet:
            continue
        da = str(d.get('address', '')).upper()
        head = da.split(':')[0]
        if a == da or a == head or da.startswith(a + ':'):
            if types is None or d.get('type') in types:
                return d
    return None


def check_assert(diffs, sheet, addr, kind, types=None, rule_expect='same'):
    """kind: 'DIFF'(真差异格) 'SAME'(无真差异格) 'PRESENT'(仅需存在) 'NODIFF'(必须无差)"""
    d = diff_at(diffs, sheet, addr, types)
    if kind == 'NODIFF':
        if d is None:
            return None, 'OK'
        return 'FAIL', '不应报差但报 %s@%s: %s' % (d.get('type'), addr, str(d.get('desc', ''))[:70])
    if kind == 'PRESENT':
        return (None, 'OK') if d is not None else ('FAIL', '缺少期望差异 @%s' % addr)
    if d is None:
        return 'FAIL', '期望 %s 但无差异 @%s' % (kind, addr)
    if kind == 'DIFF':
        if rule_expect == 'same':
            if d.get('rule_pass') is True:
                return 'FAIL', '%s@%s 期望未豁免，但被规则豁免(rule_pass=True)' % (d.get('type'), addr)
            return None, 'OK'
        if d.get('rule_pass') is not True:
            return 'FAIL', '%s@%s 期望豁免，但 rule_pass=%s rule_name=%s' % (d.get('type'), addr, d.get('rule_pass'), d.get('rule_name'))
        return None, 'OK'
    if kind == 'SAME':
        if rule_expect == 'same':
            if d.get('rule_pass') is not True:
                return 'FAIL', '%s@%s 无差格期望豁免，但未豁免(rule_name=%s)' % (d.get('type'), addr, d.get('rule_name'))
            return None, 'OK'
        if d.get('rule_pass') is True:
            return 'FAIL', '%s@%s 无差格期望不豁免，但被豁免' % (d.get('type'), addr)
        return None, 'OK'
    return 'FAIL', '未知 kind %s' % kind


def com_read_cells(path, sheet, addr):
    """COM 读单格快照：值/文本/数字格式/字体名/字号/行高/列宽/填充色(BGR)"""
    try:
        import pythoncom
        import win32com.client
        try:
            pythoncom.CoInitialize()
        except Exception:
            pass
        app = win32com.client.DispatchEx('Excel.Application')
        app.Visible = False
        app.DisplayAlerts = False
        try:
            wb = app.Workbooks.Open(path, ReadOnly=True)
            try:
                ws = wb.Worksheets(sheet)
                rng = ws.Range(addr)
                snap = {}
                for f in ('value', 'text', 'font_name', 'font_size', 'row_height', 'col_width'):
                    try:
                        snap[f] = getattr(rng, f)
                    except Exception:
                        pass
                try:
                    snap['numfmt'] = rng.NumberFormat
                except Exception:
                    pass
                try:
                    c = rng.Interior.Color
                    snap['fill_bgr'] = int(c) if c is not None else None
                except Exception:
                    pass
                return snap
            finally:
                wb.Close(False)
        finally:
            try:
                app.Quit()
            except Exception:
                pass
    except Exception as e:
        return {'error': str(e)}


def com_snap_equal(a, b, fields):
    diffs = []
    for f in fields:
        va = (a or {}).get(f)
        vb = (b or {}).get(f)
        if va != vb:
            diffs.append('%s: %s vs %s' % (f, va, vb))
    return (not diffs), '; '.join(diffs)


def l2_run_case(case, tmpdir):
    res = {'batch': case.get('batch', '?'), 'name': case.get('name', '?'),
           'fails': [], 'notes': [], 'com_lines': [], 'ok': True, 'skipped': []}
    try:
        data = case['build']()
    except Exception as e:
        res['ok'] = False
        res['fails'].append('构建失败: %s | %s' % (e, traceback.format_exc()[:200]))
        return res
    try:
        old, new = l2_save(data, res['name'], tmpdir)
    except Exception as e:
        res['ok'] = False
        res['fails'].append('保存失败: %s | %s' % (e, traceback.format_exc()[:200]))
        return res
    project = None
    if data.get('rules'):
        project = l2_make_project(data['rules'])
    copt = dict(DEFAULT_CHECK_OPTIONS)
    copt.update((data.get('opts') or {}) or {})
    cmp = OpenpyxlComparer(old, new, log_callback=lambda m: None,
                           progress_callback=lambda v, s=None: None,
                           progress_mode_fn=lambda m: None,
                           check_options=copt,
                           check_project=project,
                           stop_event=threading.Event())
    try:
        cmp.run()
    except Exception as e:
        res['ok'] = False
        res['fails'].append('引擎异常: %s' % e)
        return res
    if project is not None:
        try:
            cmp._apply_rule_filter(cmp.diffs, cmp.old_wb_ref, cmp.new_wb_ref)
        except Exception as e:
            res['fails'].append('规则过滤异常: %s' % e)
            res['ok'] = False
    rule_expect = 'same'
    for (ds, checks) in (data.get('rules') or []):
        for (t, e) in checks:
            rule_expect = e
            break
        break
    for (sh, ad, kind, types) in (data.get('asserts') or []):
        st, msg = check_assert(cmp.diffs, sh, ad, kind, types, rule_expect=rule_expect)
        if st:
            res['fails'].append('%s@%s: %s' % (ad, sh, msg))
            res['ok'] = False
    for com in (data.get('com') or []):
        kind2 = com.get('kind', 'observe')
        fields = com.get('fields', ['text'])
        a = com_read_cells(old, com.get('old_sheet', 'Sheet1'), com.get('old_addr'))
        b = com_read_cells(new, com.get('new_sheet', 'Sheet1'), com.get('new_addr'))
        if (isinstance(a, dict) and a.get('error')) or (isinstance(b, dict) and b.get('error')):
            err = (a or {}).get('error') or (b or {}).get('error')
            res['com_lines'].append('[SKIP] %s: COM %s' % (com.get('label', '?'), err))
            res['skipped'].append(com.get('label', '?'))
            continue
        eq, desc = com_snap_equal(a, b, fields)
        prog_diff = bool(com.get('prog_diff'))
        if kind2 == 'fp':
            verdict = 'FP候选(程序报差,COM相同)' if eq else 'consistent'
        elif kind2 == 'fn':
            verdict = 'FN候选(程序无差,COM不同)' if not eq else 'consistent'
        else:
            verdict = '观察(%s)' % ('COM一致' if eq else ('COM不同: ' + (desc[:80] if desc else '')))
        res['com_lines'].append('[%s] %s | 程序: %s | COM字段: %s' % (
            verdict, com.get('label', '?'), '报差' if prog_diff else '无差', desc if desc else '一致'))
    return res

def run_l2(tmpdir, out_dir, com_ok=True):
    L = []
    L.append('===== L2 全场景验证报告 =====')
    L.append('时间: %s | 版本: %s | COM可用: %s' % (time.strftime('%Y-%m-%d %H:%M:%S'),
             getattr(main, 'VERSION', '?'), '是' if com_ok else '否(COM段将SKIP)'))
    L.append('')
    results = []
    for case in suite():
        results.append(l2_run_case(case, tmpdir))
    per_batch = {}
    for r in results:
        per_batch.setdefault(r['batch'], [0, 0])
        per_batch[r['batch']][0] += 1
        if r['ok']:
            per_batch[r['batch']][1] += 1
    all_ok = sum(v[1] for v in per_batch.values())
    total = sum(v[0] for v in per_batch.values())
    for batch in sorted(per_batch):
        b = per_batch[batch]
        L.append('【%s】%d/%d 通过' % (batch, b[1], b[0]))
    L.append('')
    L.append('---- 失败明细（硬断言 ❌）----')
    fails = [r for r in results if not r['ok']]
    if not fails:
        L.append('（无失败）')
    for r in fails:
        L.append('❌ [%s] %s' % (r['batch'], r['name']))
        for f in r['fails']:
            L.append('   ' + f)
    L.append('')
    L.append('---- COM 对照观察点 ----')
    com_lines = []
    for r in results:
        com_lines.extend(r['com_lines'])
    if not com_lines:
        L.append('（无 COM 对照用例或全部 SKIP）')
    for cl in com_lines:
        L.append('  ' + cl)
    L.append('')
    L.append('汇总: 硬断言 %d/%d 通过（全场景用例）' % (all_ok, total))
    L.append('')
    L.append('说明: 所有案例的原始对照文件保留在 cases_gen 目录（l2_*_old/new.xlsx），失败项可在 Excel 中人工复核。')
    return L


# ============================================================
# 四、B1 单项矩阵
# ============================================================
def _b1_normal_build():
    old = Workbook(); ow = old.active; ow.title = 'Sheet1'
    new = Workbook(); nw = new.active; nw.title = 'Sheet1'
    for ws in (ow, nw):
        ws['A1'] = 'L2ANCHOR'
    for i in range(13):
        r = 2 + i
        ow.cell(r, 1, 100 + i); nw.cell(r, 1, 100 + i)
        if i == 8:
            nw.merge_cells('A%d:B%d' % (r, r))
            continue
        ow.cell(r, 2, 200 + i); nw.cell(r, 2, 200 + i)
        ow.cell(r, 3, 300 + i); nw.cell(r, 3, 300 + i)
        ow.cell(r, 4, 400 + i); nw.cell(r, 4, 400 + i)
        if i == 0:
            nw.cell(r, 2, 2000 + i)
            ow.cell(r, 3, 0.30000000000000004); nw.cell(r, 3, 0.3)
            ow.cell(r, 4, '=1+1'); nw.cell(r, 4, 2)
        elif i == 1:
            ow.cell(r, 2, '=1+1'); nw.cell(r, 2, '=2+2')
            ow.cell(r, 3, '=A1'); nw.cell(r, 3, '=a1')
            ow.cell(r, 4, '=SUM(A1:A1)'); nw.cell(r, 4, '=SUM(B1:B1)')
        elif i == 2:
            if HAS_RICH:
                ow.cell(r, 2, 'AB')
                nw.cell(r, 2, CellRichText(TextBlock(InlineFont(b=True), 'AB')))
                ow.cell(r, 3, CellRichText(TextBlock(InlineFont(rFont='微软雅黑'), 'AB')))
                nw.cell(r, 3, CellRichText([TextBlock(InlineFont(rFont='微软雅黑'), 'A'),
                                            TextBlock(InlineFont(rFont='微软雅黑'), 'B')]))
        elif i == 3:
            ow.cell(r, 2).font = Font(name='宋体')
            nw.cell(r, 2).font = Font(name='微软雅黑')
            ow.cell(r, 3).font = Font(name='微软雅黑')
            nw.cell(r, 3).font = Font(name='Microsoft YaHei')
        elif i == 4:
            ow.cell(r, 2).fill = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
            nw.cell(r, 2).fill = PatternFill(fill_type='solid', start_color='0000FF', end_color='0000FF')
            ow.cell(r, 3).fill = PatternFill(fill_type='solid', start_color=Color(theme=5), end_color=Color(theme=5))
            nw.cell(r, 3).fill = PatternFill(fill_type='solid', start_color='FF4472C4', end_color='FF4472C4')
            ow.cell(r, 4).fill = PatternFill(fill_type='solid', start_color=Color(theme=5, tint=-0.25), end_color=Color(theme=5, tint=-0.25))
            nw.cell(r, 4).fill = PatternFill(fill_type='solid', start_color='FF2F5597', end_color='FF2F5597')
        elif i == 5:
            thin = Side(style='thin', color='000000')
            nw.cell(r, 2).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        elif i == 6:
            nw.cell(r, 2).alignment = Alignment(horizontal='center')
        elif i == 7:
            ow.cell(r, 2).number_format = '0'
            nw.cell(r, 2).number_format = '0.00'
            ow.cell(r, 4, 0.25); nw.cell(r, 4, 0.25)
            ow.cell(r, 4).number_format = '0'
            nw.cell(r, 4).number_format = '0.00'
        elif i == 9:
            nw.row_dimensions[r].height = 30
        elif i == 10:
            nw.column_dimensions['B'].width = 25
        elif i == 11:
            if HAS_PIL:
                _fpx = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
                _pngx = _fpx.name; _fpx.close()
                PILImage.new('RGB', (10, 10), (255, 0, 0)).save(_pngx, format='PNG')
                nw.add_image(XLImage(_pngx), 'B%d' % r)
        elif i == 12:
            if CellIsRule is not None:
                nw.conditional_formatting.add('B%d' % r, CellIsRule(operator='greaterThan', formula=['5'],
                                                                    fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')))
    return old, new


def _b1_normal_asserts():
    A = []
    for i in range(13):
        r = 2 + i
        A.append(('Sheet1', 'A%d' % r, 'NODIFF', None))
        if i == 8:
            A.append(('Sheet1', 'A%d' % r, 'DIFF', ['合并新增']))
            A.append(('Sheet1', 'B%d' % r, 'NODIFF', None))
        elif i == 9:
            A.append(('Sheet1', 'A%d' % r, 'DIFF', ['行高变化']))
        elif i == 10:
            A.append(('Sheet1', 'B1', 'DIFF', ['列宽变化']))
        elif i == 11:
            if HAS_PIL:
                A.append(('Sheet1', 'B%d' % r, 'DIFF', ['图片新增', '图片变动', '图片尺寸变化']))
            else:
                A.append(('Sheet1', 'B%d' % r, 'NODIFF', None))
        elif i == 12:
            A.append(('Sheet1', 'B%d' % r, 'DIFF', ['条件格式新增', '条件格式删除', '条件格式变化']))
        else:
            A.append(('Sheet1', 'B%d' % r, 'DIFF', None))
            A.append(('Sheet1', 'C%d' % r, 'NODIFF', None))
            if i in (0, 7):
                A.append(('Sheet1', 'D%d' % r, 'DIFF', None))
            else:
                A.append(('Sheet1', 'D%d' % r, 'NODIFF', None))
    return A


def _b1_shift_build():
    old = Workbook(); ow = old.active; ow.title = 'Sheet1'
    new = Workbook(); nw = new.active; nw.title = 'Sheet1'
    for ws in (ow, nw):
        ws['A1'] = 'L2ANCHOR'; ws['B1'] = 'H2'; ws['C1'] = 'H3'; ws['D1'] = 'H4'
    for i in range(13):
        r = 2 + i
        ow.cell(r, 1, 100 + i); nw.cell(r, 3, 100 + i)
        if i == 8:
            nw.merge_cells('C%d:D%d' % (r, r))
            continue
        ow.cell(r, 2, 200 + i); nw.cell(r, 4, 200 + i)
        ow.cell(r, 3, 300 + i); nw.cell(r, 5, 300 + i)
        ow.cell(r, 4, 400 + i); nw.cell(r, 6, 400 + i)
        if i == 0:
            nw.cell(r, 4, 2000 + i)
            ow.cell(r, 3, 0.30000000000000004); nw.cell(r, 5, 0.3)
            ow.cell(r, 4, '=1+1'); nw.cell(r, 6, 2)
        elif i == 1:
            ow.cell(r, 2, '=1+1'); nw.cell(r, 4, '=2+2')
            ow.cell(r, 3, '=A1'); nw.cell(r, 5, '=a1')
            ow.cell(r, 4, '=SUM(A1:A1)'); nw.cell(r, 6, '=SUM(B1:B1)')
        elif i == 2:
            if HAS_RICH:
                ow.cell(r, 2, 'AB')
                nw.cell(r, 4, CellRichText(TextBlock(InlineFont(b=True), 'AB')))
                ow.cell(r, 3, CellRichText(TextBlock(InlineFont(rFont='微软雅黑'), 'AB')))
                nw.cell(r, 5, CellRichText([TextBlock(InlineFont(rFont='微软雅黑'), 'A'),
                                            TextBlock(InlineFont(rFont='微软雅黑'), 'B')]))
        elif i == 3:
            ow.cell(r, 2).font = Font(name='宋体')
            nw.cell(r, 4).font = Font(name='微软雅黑')
            ow.cell(r, 3).font = Font(name='微软雅黑')
            nw.cell(r, 5).font = Font(name='Microsoft YaHei')
        elif i == 4:
            ow.cell(r, 2).fill = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
            nw.cell(r, 4).fill = PatternFill(fill_type='solid', start_color='0000FF', end_color='0000FF')
            ow.cell(r, 3).fill = PatternFill(fill_type='solid', start_color=Color(theme=5), end_color=Color(theme=5))
            nw.cell(r, 5).fill = PatternFill(fill_type='solid', start_color='FF4472C4', end_color='FF4472C4')
            ow.cell(r, 4).fill = PatternFill(fill_type='solid', start_color=Color(theme=5, tint=-0.25), end_color=Color(theme=5, tint=-0.25))
            nw.cell(r, 6).fill = PatternFill(fill_type='solid', start_color='FF2F5597', end_color='FF2F5597')
        elif i == 5:
            thin = Side(style='thin', color='000000')
            nw.cell(r, 4).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        elif i == 6:
            nw.cell(r, 4).alignment = Alignment(horizontal='center')
        elif i == 7:
            ow.cell(r, 2).number_format = '0'
            nw.cell(r, 4).number_format = '0.00'
            ow.cell(r, 4, 0.25); nw.cell(r, 6, 0.25)
            ow.cell(r, 4).number_format = '0'
            nw.cell(r, 6).number_format = '0.00'
        elif i == 9:
            nw.row_dimensions[r].height = 30
        elif i == 10:
            nw.column_dimensions['B'].width = 30
            nw.column_dimensions['C'].width = 25
        elif i == 11:
            if HAS_PIL:
                _fpx = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
                _pngx = _fpx.name; _fpx.close()
                PILImage.new('RGB', (10, 10), (255, 0, 0)).save(_pngx, format='PNG')
                nw.add_image(XLImage(_pngx), 'C%d' % r)
        elif i == 12:
            if CellIsRule is not None:
                nw.conditional_formatting.add('D%d' % r, CellIsRule(operator='greaterThan', formula=['5'],
                                                                    fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')))
    return old, new


def _b1_shift_asserts():
    A = []
    for i in range(13):
        r = 2 + i
        if i == 8:
            A.append(('Sheet1', 'A%d' % r, 'DIFF', ['单元格删除']))
            A.append(('Sheet1', 'C%d' % r, 'DIFF', ['合并新增']))
        elif i == 9:
            A.append(('Sheet1', 'A%d' % r, 'DIFF', ['单元格删除']))
            A.append(('Sheet1', 'A%d' % r, 'PRESENT', ['行高变化']))
        elif i == 10:
            A.append(('Sheet1', 'C1', 'DIFF', ['列宽变化']))
            A.append(('Sheet1', 'B1', 'PRESENT', ['列宽变化']))
            A.append(('Sheet1', 'B%d' % r, 'SAME', ['单元格删除']))
        elif i == 11:
            A.append(('Sheet1', 'A%d' % r, 'DIFF', ['单元格删除']))
            if HAS_PIL:
                A.append(('Sheet1', 'C%d' % r, 'DIFF', ['图片新增', '图片变动', '图片尺寸变化']))
            else:
                A.append(('Sheet1', 'C%d' % r, 'NODIFF', None))
        elif i == 12:
            A.append(('Sheet1', 'A%d' % r, 'SAME', ['单元格删除']))
            A.append(('Sheet1', 'D%d' % r, 'DIFF', ['条件格式新增', '条件格式删除', '条件格式变化']))
        else:
            A.append(('Sheet1', 'A%d' % r, 'SAME', ['单元格删除']))
            A.append(('Sheet1', 'B%d' % r, 'DIFF', ['单元格删除']))
            A.append(('Sheet1', 'C%d' % r, 'SAME', ['内容变化']))
            if i in (0, 7):
                A.append(('Sheet1', 'D%d' % r, 'DIFF', ['内容变化']))
            else:
                A.append(('Sheet1', 'D%d' % r, 'SAME', ['内容变化']))
    return A


def _b1_cases():
    out = []
    layouts = [('普通', _b1_normal_build(), _b1_normal_asserts()),
               ('shift', _b1_shift_build(), _b1_shift_asserts())]
    for layout, (o, n), asserts in layouts:
        for exp in ('same', 'different'):
            if layout == '普通':
                if exp == 'same':
                    rules = [(_range_ds('Sheet1', 'L2ANCHOR', 1, 0, 13, 4), [(t, 'same') for t in CHECK_TYPES])]
                else:
                    rules = [(_range_ds('Sheet1', 'L2ANCHOR', 1 + i, 0, 1, 4), [(t, 'different')])
                             for i, t in enumerate(CHECK_TYPES)]
            else:
                rules = [(_shift_ds('Sheet1', 'L2ANCHOR', 2, str(2 + i)), [(t, exp)])
                         for i, t in enumerate(CHECK_TYPES)]
            out.append({'batch': 'B1', 'name': '%s规则·期望%s' % (layout, exp),
                        'build': (lambda o=o, n=n, rules=rules, asserts=asserts:
                                  {'old': o, 'new': n, 'rules': rules, 'asserts': asserts})})
    return out

# ============================================================
# 五、B2 干扰矩阵
# ============================================================
def _b2_build(diff_types, rule_checks, expect, opts=None, make_diff=True):
    old = Workbook(); ow = old.active; ow.title = 'Sheet1'
    new = Workbook(); nw = new.active; nw.title = 'Sheet1'
    for ws in (ow, nw):
        ws['A1'] = 'L2ANCHOR'
        ws['A2'] = 10; ws['B2'] = 'keep'
    if make_diff:
        _fill_diff(ow, nw, 2, 1, diff_types)
    rules = [(_range_ds('Sheet1', 'L2ANCHOR', 1, 0, 2, 2), [(t, expect) for t in rule_checks])]
    return {'old': old, 'new': new, 'rules': rules, 'opts': opts or {},
            'asserts': [('Sheet1', 'A2', 'DIFF', None), ('Sheet1', 'B2', 'NODIFF', None)]}


def _b2_cases():
    from itertools import combinations
    out = []
    avail = [t for t in CHECK_TYPES if t in ('value', 'formula', 'rich_text', 'font', 'fill', 'border',
                                             'alignment', 'number_format', 'merged_cells', 'row_height',
                                             'col_width', 'images', 'conditional_format')]
    if not HAS_RICH:
        avail = [t for t in avail if t != 'rich_text']
    if not HAS_PIL:
        avail = [t for t in avail if t != 'images']
    for i, j in combinations(range(len(avail)), 2):
        t1, t2 = avail[i], avail[j]
        if t1 == 'col_width':
            t1, t2 = t2, t1
        cs = [t1, t2]
        out.append({'batch': 'B2', 'name': '两两 %s+%s·same' % (t1, t2),
                    'build': (lambda cs=cs: _b2_build([cs[0]], cs, 'same'))})
    triples = [
        ('value', 'font', 'fill'), ('value', 'border', 'alignment'),
        ('value', 'number_format', 'conditional_format'), ('value', 'rich_text', 'images'),
        ('formula', 'font', 'fill'), ('formula', 'border', 'alignment'),
        ('formula', 'number_format', 'conditional_format'), ('formula', 'rich_text', 'images'),
        ('font', 'fill', 'border'), ('alignment', 'number_format', 'conditional_format'),
        ('rich_text', 'images', 'merged_cells'), ('value', 'merged_cells', 'conditional_format'),
        ('formula', 'merged_cells', 'row_height'), ('value', 'row_height', 'col_width'),
        ('font', 'alignment', 'conditional_format'), ('fill', 'number_format', 'rich_text'),
        ('formula', 'images', 'conditional_format'), ('value', 'font', 'conditional_format'),
    ]
    triples = [tp for tp in triples if all(t in avail for t in tp)]
    for tp in triples:
        out.append({'batch': 'B2', 'name': '单格多差 %s·same' % '+'.join(tp),
                    'build': (lambda tp=tp: _b2_build(list(tp), list(tp), 'same'))})
        out.append({'batch': 'B2', 'name': '单格多差 %s·different' % '+'.join(tp),
                    'build': (lambda tp=tp: _b2_build(list(tp), list(tp), 'different'))})
    return out


def _b2_full13_build():
    old = Workbook(); ow = old.active; ow.title = 'Sheet1'
    new = Workbook(); nw = new.active; nw.title = 'Sheet1'
    for ws in (ow, nw):
        ws['A1'] = 'L2ANCHOR'; ws['B1'] = 'keep'
    # 1) 先写 A 列基准值
    for i in range(13):
        r = 2 + i
        ow.cell(r, 1, 100 + i); nw.cell(r, 1, 100 + i)
    # 2) 写 B 列基准值（跳过 i=8/9，因为 B10:B11 会被合并，B11 不可写）
    for i in range(13):
        r = 2 + i
        if i == 8 or i == 9:
            continue
        ow.cell(r, 2, 200 + i); nw.cell(r, 2, 200 + i)
    # 3) 逐个制造类型差异（merged 用 B 列、跨行但不写合并区外的值）
    for i in range(13):
        r = 2 + i
        t = CHECK_TYPES[i]
        if i == 8:
            nw.merge_cells('B%d:B%d' % (r, r + 1))
            continue
        if i == 10:
            nw.column_dimensions['B'].width = 25
            continue
        _fill_diff(ow, nw, r, 1, [t])
    return old, new


def _b2_full13_asserts(kind='DIFF'):
    A = []
    for i in range(13):
        r = 2 + i
        if i == 8:
            if kind == 'NODIFF':
                A.append(('Sheet1', 'A%d' % r, 'NODIFF', None))
            continue
        if i == 10:
            # 列宽差异在 B1（列头），A12 本身无单元级差异
            if kind == 'NODIFF':
                A.append(('Sheet1', 'A%d' % r, 'NODIFF', None))
            continue
        if i == 11 and not HAS_PIL:
            A.append(('Sheet1', 'A%d' % r, 'NODIFF', None))
            continue
        A.append(('Sheet1', 'A%d' % r, kind, None))
    if kind == 'DIFF':
        A.append(('Sheet1', 'B10', 'DIFF', ['合并新增']))
        A.append(('Sheet1', 'B1', 'DIFF', ['列宽变化']))
    else:
        A.append(('Sheet1', 'B10', 'NODIFF', None))
        A.append(('Sheet1', 'B1', 'NODIFF', None))
    return A


def _b2_full_cases():
    out = []
    o, n = _b2_full13_build()
    rules = [(_range_ds('Sheet1', 'L2ANCHOR', 1, 0, 13, 2), [(t, 'same') for t in CHECK_TYPES])]
    out.append({'batch': 'B2', 'name': '全开13·每格单项差·same',
                'build': (lambda o=o, n=n, rules=rules:
                          {'old': o, 'new': n, 'rules': rules, 'asserts': _b2_full13_asserts('DIFF')})})
    return out


# ============================================================
# 六、B3 全局交叉
# ============================================================
def _b3_cases():
    out = []
    ds = _range_ds('Sheet1', 'L2ANCHOR', 1, 0, 2, 2)
    combos = [
        ('B3-1 全局开·规则value same', [('value', 'same')], {}, 'A2', 'DIFF'),
        ('B3-2 全局开·规则value different', [('value', 'different')], {}, 'A2', 'DIFF'),
        ('B3-3 全局关·规则value same', [('value', 'same')], {'value': False}, 'A2', 'NODIFF'),
        ('B3-4 全局关·规则value different', [('value', 'different')], {'value': False}, 'A2', 'NODIFF'),
        ('B3-5 全局开·规则仅font same', [('font', 'same')], {}, 'A2', 'SAME'),
        ('B3-6 全局开·空检查规则', [], {}, 'A2', 'SAME'),
    ]
    for name, checks, opts, addr, kind in combos:
        exp = checks[0][1] if checks else 'same'
        rule_types = [t for t, e in checks]
        out.append({'batch': 'B3', 'name': name,
                    'build': (lambda rule_types=rule_types, exp=exp, opts=opts, addr=addr, kind=kind:
                              dict(_b2_build(['value'], rule_types, exp, opts),
                                   **{'asserts': [('Sheet1', addr, kind, None)]}))})
    for name, checks in (('B3-7 无差+规则same', [('value', 'same')]),
                         ('B3-8 无差+规则different', [('value', 'different')])):
        out.append({'batch': 'B3', 'name': name,
                    'build': (lambda checks=checks:
                              dict(_b2_build([], [t for t, e in checks], checks[0][1]),
                                   **{'asserts': [('Sheet1', 'A2', 'NODIFF', None)]}))})
    o, n = _b2_full13_build()
    rules = [(_range_ds('Sheet1', 'L2ANCHOR', 1, 0, 13, 2), [(t, 'different') for t in CHECK_TYPES])]
    allopt = {t: False for t in CHECK_TYPES}
    out.append({'batch': 'B3', 'name': 'B3-B 全局关13项+规则13开different',
                'build': (lambda o=o, n=n, rules=rules, allopt=allopt:
                          {'old': o, 'new': n, 'rules': rules, 'opts': allopt,
                           'asserts': _b2_full13_asserts('NODIFF')})})
    out.append({'batch': 'B3', 'name': 'B3-C 无规则·值差',
                'build': (lambda: dict(_b2_build(['value'], ['value'], 'same'),
                                       **{'rules': None, 'asserts': [('Sheet1', 'A2', 'PRESENT', None)]}))})
    return out

# ============================================================
# 七、B4 COM 分歧 12 case
# ============================================================
def _b4_base():
    """B4/B5 共用基础文件对：A1=锚点文字, B1=5"""
    old = Workbook(); ow = old.active; ow.title = 'Sheet1'
    new = Workbook(); nw = new.active; nw.title = 'Sheet1'
    for ws in (ow, nw):
        ws['A1'] = 'L2ANCHOR'
        ws['B1'] = 5
    return old, new


def _b4_cases():
    out = []

    def _mk(prog_diff, fields, kind='observe', label='', old_addr='A2', new_addr='A2'):
        return [{'kind': kind, 'label': label or old_addr, 'old_sheet': 'Sheet1', 'old_addr': old_addr,
                 'new_sheet': 'Sheet1', 'new_addr': new_addr, 'fields': fields, 'prog_diff': prog_diff}]

    def c(name, make, asserts, com):
        def build():
            ow, nw = _b4_base()
            make(ow, nw)
            return {'old': ow, 'new': nw, 'asserts': asserts, 'com': com}
        out.append({'batch': 'B4', 'name': name, 'build': build})

    def m1(ow, nw):
        ow['A2'] = '=1+1'; nw['A2'] = '=2'
    c('B4-1 公式缓存值', m1, [('Sheet1', 'A2', 'PRESENT', None)],
      _mk(True, ['text'], 'observe', '公式显示值'))

    def m2(ow, nw):
        ow['A2'] = 0.30000000000000004; nw['A2'] = 0.3
    c('B4-2 浮点尾差', m2, [('Sheet1', 'A2', 'NODIFF', None)],
      _mk(False, ['text'], 'observe', '尾差显示'))

    def m3(ow, nw):
        ow['A2'] = 0.25; nw['A2'] = 0.25
        ow['A2'].number_format = '0'; nw['A2'].number_format = '0.00'
    c('B4-3 数字格式显示', m3, [('Sheet1', 'A2', 'PRESENT', ['数字格式变化'])],
      _mk(True, ['text'], 'observe', '格式显示'))

    def m4(ow, nw):
        ow['A2'] = 'x'; nw['A2'] = 'x'
        ow['A2'].font = Font(name='微软雅黑')
        nw['A2'].font = Font(name='Microsoft YaHei')
    c('B4-4 字体别名等价', m4, [('Sheet1', 'A2', 'NODIFF', None)],
      _mk(False, ['font_name'], 'observe', '字体别名'))

    def m5(ow, nw):
        ow['A2'] = 10; nw['A2'] = 10
        ow['A2'].fill = PatternFill(fill_type='solid', start_color=Color(theme=5, tint=-0.25), end_color=Color(theme=5, tint=-0.25))
        nw['A2'].fill = PatternFill(fill_type='solid', start_color='FF2F5597', end_color='FF2F5597')
    c('B4-5 主题tint±1', m5, [('Sheet1', 'A2', 'NODIFF', None)],
      _mk(False, ['fill_bgr'], 'observe', 'tint色'))

    def m6(ow, nw):
        ow['A2'] = 10; nw['A2'] = 10
        ow['A2'].fill = PatternFill(fill_type='solid', start_color=Color(indexed=23), end_color=Color(indexed=23))
        nw['A2'].fill = PatternFill(fill_type='solid', start_color='FF808080', end_color='FF808080')
    c('B4-6 indexed等价', m6, [('Sheet1', 'A2', 'NODIFF', None)],
      _mk(False, ['fill_bgr'], 'observe', 'indexed色'))

    def m7(ow, nw):
        nw.row_dimensions[2].height = 15
    c('B4-7 行高默认vs15', m7, [('Sheet1', 'A2', 'PRESENT', ['行高变化'])],
      _mk(True, ['row_height'], 'fp', '行高15'))

    def m8(ow, nw):
        nw.column_dimensions['B'].width = 8.43
    c('B4-8 列宽默认vs8.43', m8, [('Sheet1', 'B1', 'PRESENT', ['列宽变化'])],
      _mk(True, ['col_width'], 'fp', '列宽8.43', 'B1', 'B1'))

    def m9(ow, nw):
        ow['A2'] = 'AB'
        nw['A2'] = CellRichText(TextBlock(InlineFont(b=True), 'AB'))
    c('B4-9 富文本样式', m9, [('Sheet1', 'A2', 'PRESENT', None)],
      _mk(True, ['text'], 'observe', '富文本'))

    def m10(ow, nw):
        ow['A2'] = 10; nw['A2'] = 10
        if CellIsRule is not None:
            nw.conditional_formatting.add('A2', CellIsRule(operator='greaterThan', formula=['5'],
                                                           fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')))
    c('B4-10 条件格式', m10, [('Sheet1', 'A2', 'PRESENT', ['条件格式新增', '条件格式删除', '条件格式变化'])],
      _mk(True, ['text'], 'observe', '条件格式'))

    def m11(ow, nw):
        ow['A2'] = 10; nw['A2'] = 10
        ow['B2'] = 10; nw['B2'] = 10
        ow.merge_cells('A2:B2')
    c('B4-11 合并删除', m11, [('Sheet1', 'A2', 'PRESENT', ['合并删除'])],
      _mk(True, ['text'], 'observe', '合并'))

    def m12(ow, nw):
        ow['A2'] = '=B1'; nw['A2'] = '=b1'
    c('B4-12 公式大小写', m12, [('Sheet1', 'A2', 'NODIFF', None)],
      _mk(False, ['text'], 'observe', '公式大小写'))

    return out


# ============================================================
# 八、B5 主题/字体等价专项
# ============================================================
def _mk_b5(prog_diff, fields, kind='observe', label='', old_addr='A2', new_addr='A2'):
    return [{'kind': kind, 'label': label or old_addr, 'old_sheet': 'Sheet1', 'old_addr': old_addr,
             'new_sheet': 'Sheet1', 'new_addr': new_addr, 'fields': fields, 'prog_diff': prog_diff}]


def _b5_cases():
    out = []

    def c(name, make, asserts, com, need_reload=False):
        def build():
            ow, nw = _b4_base()
            make(ow, nw)
            if need_reload:
                fp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
                sp = fp.name; fp.close()
                nw.save(sp)
                _set_theme_accent1(sp, 'FF0000')
                nw = load_workbook(sp)
                os.unlink(sp)
            return {'old': ow, 'new': nw, 'asserts': asserts, 'com': com}
        out.append({'batch': 'B5', 'name': name, 'build': build})

    def m1(ow, nw):
        ow['A2'] = 10; nw['A2'] = 10
        ow['A2'].fill = PatternFill(fill_type='solid', start_color=Color(theme=5), end_color=Color(theme=5))
        nw['A2'].fill = PatternFill(fill_type='solid', start_color='FF4472C4', end_color='FF4472C4')
    c('B5-1 主题色vsRGB', m1, [('Sheet1', 'A2', 'NODIFF', None)],
      _mk_b5(False, ['fill_bgr'], 'observe', '主题色'))

    def m2(ow, nw):
        ow['A2'] = 10; nw['A2'] = 10
        ow['A2'].fill = PatternFill(fill_type='solid', start_color=Color(theme=5), end_color=Color(theme=5))
        nw['A2'].fill = PatternFill(fill_type='solid', start_color=Color(theme=5), end_color=Color(theme=5))
    c('B5-2 主题表不同', m2, [('Sheet1', 'A2', 'PRESENT', ['填充变化'])],
      _mk_b5(True, ['fill_bgr'], 'observe', '跨主题表'), need_reload=True)

    def m3(ow, nw):
        ow['A2'] = 10; nw['A2'] = 10
        ow['A2'].fill = PatternFill(fill_type='solid', start_color=Color(theme=5, tint=-0.25), end_color=Color(theme=5, tint=-0.25))
        nw['A2'].fill = PatternFill(fill_type='solid', start_color='FF2F5597', end_color='FF2F5597')
    c('B5-3 tint±1', m3, [('Sheet1', 'A2', 'NODIFF', None)],
      _mk_b5(False, ['fill_bgr'], 'observe', 'tint色'))

    def m4(ow, nw):
        ow['A2'] = 10; nw['A2'] = 10
        ow['A2'].fill = PatternFill(fill_type='solid', start_color=Color(indexed=23), end_color=Color(indexed=23))
        nw['A2'].fill = PatternFill(fill_type='solid', start_color='FF808080', end_color='FF808080')
    c('B5-4 indexed等价', m4, [('Sheet1', 'A2', 'NODIFF', None)],
      _mk_b5(False, ['fill_bgr'], 'observe', 'indexed色'))

    def m5(ow, nw):
        ow['A2'] = 'x'; nw['A2'] = 'x'
        ow['A2'].font = Font(name='微软雅黑')
        nw['A2'].font = Font(name='Microsoft YaHei')
    c('B5-5 字体别名', m5, [('Sheet1', 'A2', 'NODIFF', None)],
      _mk_b5(False, ['font_name'], 'observe', '字体别名'))

    def m6(ow, nw):
        ow['A2'] = 'x'; nw['A2'] = 'x'
        ow['A2'].font = Font(scheme='minor')
        nw['A2'].font = Font(name='Calibri')
    c('B5-6 主题字体scheme', m6, [('Sheet1', 'A2', 'NODIFF', None)],
      _mk_b5(False, ['font_name'], 'observe', 'scheme字体'))

    def m7(ow, nw):
        ow['A2'] = 10; nw['A2'] = 10
        ow['A2'].font = Font(name='Calibri', size=11)
        ow['A2'].fill = PatternFill(fill_type='solid', start_color='FF4472C4', end_color='FF4472C4')
        nw['A2'].font = Font(scheme='minor', size=11)
        nw['A2'].fill = PatternFill(fill_type='solid', start_color=Color(theme=5), end_color=Color(theme=5))
    c('B5-7 组合等价', m7, [('Sheet1', 'A2', 'NODIFF', None)],
      _mk_b5(False, ['font_name', 'fill_bgr'], 'observe', '组合等价'))

    def m8(ow, nw):
        ow['A2'] = 'x'; nw['A2'] = 'x'
        ow['A2'].font = Font(name='微软雅黑')
        nw['A2'].font = Font(name='宋体')
    c('B5-8 真字体差异', m8, [('Sheet1', 'A2', 'PRESENT', ['字体变化'])],
      _mk_b5(True, ['font_name'], 'observe', '真差异'))

    return out


# ============================================================
# 九、用例汇总
# ============================================================
def suite():
    out = []
    out.extend(_b1_cases())
    out.extend(_b2_cases())
    out.extend(_b2_full_cases())
    out.extend(_b3_cases())
    out.extend(_b4_cases())
    out.extend(_b5_cases())
    return out


# ============================================================
# 十、入口
# ============================================================
def main():
    tmpdir = os.path.join(BASE, 'cases_gen')
    os.makedirs(tmpdir, exist_ok=True)
    out_dir = os.path.join(BASE, 'out')
    os.makedirs(out_dir, exist_ok=True)

    L = []
    L.append('===== 金标准回归结果 =====')
    L.append('时间: %s | 版本: %s' % (time.strftime('%Y-%m-%d %H:%M:%S'), getattr(main, 'VERSION', '?')))
    L.append('')
    passed = 0
    n = len(CASES)
    for name, builder in CASES:
        print('[回归] %s ...' % name)
        r = run_one(name, builder, tmpdir)
        if r['ok']:
            passed += 1
            L.append('✅ %s | 符合预期' % name)
        else:
            L.append('❌ %s\n   期望/问题: %s' % (name, r['detail']))
        if r.get('counts'):
            L.append('   实际检出类型: %s' % dict(r['counts']))
        if not r['ok'] and r.get('found'):
            L.append('   实际检出(前20): %s' % r['found'])
        L.append('')
    L.append('汇总: %d / %d 通过' % (passed, n))
    if passed == n:
        L.append('结论: 对比引擎核心检测逻辑正常。')
    else:
        L.append('结论: 有 %d 项不符合预期，请将本文件粘贴给 AI 分析。' % (n - passed))
    L.append('')
    L.append('说明: 案例文件保留在 cases_gen 目录，可在 Excel 中打开人工复核。')
    out_path = os.path.join(out_dir, 'RULES_VERDICT.txt')
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(L))

    com_ok = True
    try:
        import pythoncom  # noqa
        import win32com.client  # noqa
        a = win32com.client.DispatchEx('Excel.Application')
        a.Quit()
    except Exception:
        com_ok = False
    print('[L2] 开始全场景验证（约 2~4 分钟）...')
    L2 = run_l2(tmpdir, out_dir, com_ok)
    l2_path = os.path.join(out_dir, 'L2_REPORT.txt')
    with open(l2_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(L2))

    print('\n'.join(L))
    print('---')
    print('\n'.join(L2))
    print('已生成: %s' % out_path)
    print('已生成: %s' % l2_path)
    try:
        input('按回车关闭窗口...')
    except Exception:
        pass
    return 0


if __name__ == '__main__':
    sys.exit(main())

