# -*- coding: utf-8 -*-
"""
L2 全场景用例数据（供 regression.py 调用）
suite() -> list[case]；case['build']() -> dict:
  {'old','new','rules','opts','asserts','notes','com'}
  rules: [(ds, [(check_type, expect), ...]), ...]
  asserts: [(sheet, addr, kind, types_or_None)]  kind: DIFF/SAME/PRESENT/NODIFF
  com: [(dict)] kind: fp/fn/observe, fields: 比较字段, prog_diff: 程序是否报差
"""
import io, os, re, zipfile, tempfile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
try:
    from openpyxl.formatting.rule import CellIsRule
except Exception:
    CellIsRule = None
try:
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    HAS_RICH = True
except Exception:
    HAS_RICH = False
try:
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage
    HAS_PIL = True
except Exception:
    HAS_PIL = False

CHECK_TYPES = ['value', 'formula', 'rich_text', 'font', 'fill', 'border',
               'alignment', 'number_format', 'merged_cells', 'row_height',
               'col_width', 'images', 'conditional_format']


def _range_ds(sheet, anchor, r_off, c_off, r_cnt, c_cnt, name='L2'):
    return {'name': name, 'sheet': sheet, 'anchor': {'text': anchor}, 'search_in': 'all',
            'mode': 'range',
            'target': {'row_offset': r_off, 'col_offset': c_off,
                       'row_count': r_cnt, 'col_count': c_cnt}}


def _shift_ds(sheet, anchor, shift_offset, rows, name='L2'):
    return {'name': name, 'sheet': sheet, 'anchor': {'text': anchor}, 'search_in': 'all',
            'mode': 'shift',
            'header_target': {'row_offset': 0, 'col_offset': 0, 'row_count': 1, 'col_count': 4},
            'rows': rows, 'shift_offset': shift_offset}


def _fill_diff(ow, nw, r, c, kinds):
    """在普通布局 (r,c) 制造指定类型的差异（同地址）"""
    cell = '%s%d' % (get_column_letter(c), r)
    if 'value' in kinds:
        ow.cell(r, c, 10); nw.cell(r, c, 20)
    if 'formula' in kinds:
        ow.cell(r, c, '=1+1'); nw.cell(r, c, '=2+2')
    if 'rich_text' in kinds and HAS_RICH:
        ow.cell(r, c, 'AB')
        nw.cell(r, c, CellRichText(TextBlock(InlineFont(b=True), 'AB')))
    if 'font' in kinds:
        ow.cell(r, c, 'x').font = Font(name='宋体')
        nw.cell(r, c, 'x').font = Font(name='微软雅黑')
    if 'fill' in kinds:
        ow.cell(r, c).fill = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
        nw.cell(r, c).fill = PatternFill(fill_type='solid', start_color='0000FF', end_color='0000FF')
    if 'border' in kinds:
        thin = Side(style='thin', color='000000')
        nw.cell(r, c).border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if 'alignment' in kinds:
        nw.cell(r, c).alignment = Alignment(horizontal='center')
    if 'number_format' in kinds:
        ow.cell(r, c).number_format = '0'
        nw.cell(r, c).number_format = '0.00'
    if 'merged_cells' in kinds:
        nw.merge_cells('%s%d:%s%d' % (get_column_letter(c), r, get_column_letter(c), r + 1))
    if 'row_height' in kinds:
        nw.row_dimensions[r].height = 30
    if 'col_width' in kinds:
        nw.column_dimensions[get_column_letter(c)].width = 25
    if 'images' in kinds and HAS_PIL:
        buf = io.BytesIO()
        PILImage.new('RGB', (10, 10), (255, 0, 0)).save(buf, format='PNG')
        buf.seek(0)
        nw.add_image(XLImage(buf), cell)
    if 'conditional_format' in kinds and CellIsRule is not None:
        nw.conditional_formatting.add(cell, CellIsRule(operator='greaterThan', formula=['5'],
                                                       fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')))


def _set_theme_accent1(path, hex6):
    """替换 xlsx 主题表 accent1 颜色（zip 重写 theme1.xml）"""
    tmp = path + '.tmp'
    zin = zipfile.ZipFile(path, 'r')
    zout = zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED)
    for item in zin.infolist():
        data = zin.read(item.filename)
        if item.filename == 'xl/theme/theme1.xml':
            try:
                s = data.decode('utf-8')
                s2 = re.sub(r'(<a:accent1>.*?<a:srgbClr val=")[0-9A-Fa-f]{6}(?="/>)',
                            r'\g<1>' + hex6, s, flags=re.S)
                data = s2.encode('utf-8')
            except Exception:
                pass
        zout.writestr(item, data)
    zin.close()
    zout.close()
    os.replace(tmp, path)


# ------------------------------------------------------------
# B1 单项矩阵
# ------------------------------------------------------------
def _b1_normal_build():
    old = Workbook(); ow = old.active; ow.title = 'Sheet1'
    new = Workbook(); nw = new.active; nw.title = 'Sheet1'
    for ws in (ow, nw):
        ws['A1'] = 'L2ANCHOR'
    for i in range(13):
        r = 2 + i
        ow.cell(r, 1, 100 + i); nw.cell(r, 1, 100 + i)
        if i == 8:
            nw.merge_cells('A%d:B%d' % (r, r))
            continue
        ow.cell(r, 2, 200 + i); nw.cell(r, 2, 200 + i)
        ow.cell(r, 3, 300 + i); nw.cell(r, 3, 300 + i)
        ow.cell(r, 4, 400 + i); nw.cell(r, 4, 400 + i)
        if i == 0:
            nw.cell(r, 2, 2000 + i)
            ow.cell(r, 3, 0.30000000000000004); nw.cell(r, 3, 0.3)
            ow.cell(r, 4, '=1+1'); nw.cell(r, 4, 2)
        elif i == 1:
            ow.cell(r, 2, '=1+1'); nw.cell(r, 2, '=2+2')
            ow.cell(r, 3, '=A1'); nw.cell(r, 3, '=a1')
            ow.cell(r, 4, '=SUM(A1:A1)'); nw.cell(r, 4, '=SUM(B1:B1)')
        elif i == 2:
            if HAS_RICH:
                ow.cell(r, 2, 'AB')
                nw.cell(r, 2, CellRichText(TextBlock(InlineFont(b=True), 'AB')))
                ow.cell(r, 3, CellRichText(TextBlock(InlineFont(rFont='微软雅黑'), 'AB')))
                nw.cell(r, 3, CellRichText([TextBlock(InlineFont(rFont='微软雅黑'), 'A'),
                                            TextBlock(InlineFont(rFont='微软雅黑'), 'B')]))
        elif i == 3:
            ow.cell(r, 2).font = Font(name='宋体')
            nw.cell(r, 2).font = Font(name='微软雅黑')
            ow.cell(r, 3).font = Font(name='微软雅黑')
            nw.cell(r, 3).font = Font(name='Microsoft YaHei')
        elif i == 4:
            ow.cell(r, 2).fill = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
            nw.cell(r, 2).fill = PatternFill(fill_type='solid', start_color='0000FF', end_color='0000FF')
            ow.cell(r, 3).fill = PatternFill(fill_type='solid', start_color=Color(theme=5), end_color=Color(theme=5))
            nw.cell(r, 3).fill = PatternFill(fill_type='solid', start_color='FF4472C4', end_color='FF4472C4')
            ow.cell(r, 4).fill = PatternFill(fill_type='solid', start_color=Color(theme=5, tint=-0.25), end_color=Color(theme=5, tint=-0.25))
            nw.cell(r, 4).fill = PatternFill(fill_type='solid', start_color='FF2F5597', end_color='FF2F5597')
        elif i == 5:
            thin = Side(style='thin', color='000000')
            nw.cell(r, 2).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        elif i == 6:
            nw.cell(r, 2).alignment = Alignment(horizontal='center')
        elif i == 7:
            ow.cell(r, 2).number_format = '0'
            nw.cell(r, 2).number_format = '0.00'
            ow.cell(r, 4, 0.25); nw.cell(r, 4, 0.25)
            ow.cell(r, 4).number_format = '0'
            nw.cell(r, 4).number_format = '0.00'
        elif i == 9:
            nw.row_dimensions[r].height = 30
        elif i == 10:
            nw.column_dimensions['B'].width = 25
        elif i == 11:
            if HAS_PIL:
                buf = io.BytesIO()
                PILImage.new('RGB', (10, 10), (255, 0, 0)).save(buf, format='PNG')
                buf.seek(0)
                nw.add_image(XLImage(buf), 'B%d' % r)
        elif i == 12:
            if CellIsRule is not None:
                nw.conditional_formatting.add('B%d' % r, CellIsRule(operator='greaterThan', formula=['5'],
                                                                    fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')))
    return old, new


def _b1_normal_asserts():
    A = []
    for i in range(13):
        r = 2 + i
        A.append(('Sheet1', 'A%d' % r, 'NODIFF', None))
        if i == 8:
            A.append(('Sheet1', 'A%d' % r, 'DIFF', ['合并新增']))
            A.append(('Sheet1', 'B%d' % r, 'NODIFF', None))
        elif i == 9:
            A.append(('Sheet1', 'A%d' % r, 'DIFF', ['行高变化']))
        elif i == 10:
            A.append(('Sheet1', 'B1', 'DIFF', ['列宽变化']))
        elif i == 11:
            if HAS_PIL:
                A.append(('Sheet1', 'B%d' % r, 'DIFF', ['图片新增', '图片变动', '图片尺寸变化']))
            else:
                A.append(('Sheet1', 'B%d' % r, 'NODIFF', None))
        elif i == 12:
            A.append(('Sheet1', 'B%d' % r, 'DIFF', ['条件格式新增', '条件格式删除', '条件格式变化']))
        else:
            A.append(('Sheet1', 'B%d' % r, 'DIFF', None))
            A.append(('Sheet1', 'C%d' % r, 'NODIFF', None))
            if i in (0, 7):
                A.append(('Sheet1', 'D%d' % r, 'DIFF', None))
            else:
                A.append(('Sheet1', 'D%d' % r, 'NODIFF', None))
    return A


def _b1_shift_build():
    old = Workbook(); ow = old.active; ow.title = 'Sheet1'
    new = Workbook(); nw = new.active; nw.title = 'Sheet1'
    for ws in (ow, nw):
        ws['A1'] = 'L2ANCHOR'; ws['B1'] = 'H2'; ws['C1'] = 'H3'; ws['D1'] = 'H4'
    for i in range(13):
        r = 2 + i
        ow.cell(r, 1, 100 + i); nw.cell(r, 3, 100 + i)
        if i == 8:
            nw.merge_cells('C%d:D%d' % (r, r))
            continue
        ow.cell(r, 2, 200 + i); nw.cell(r, 4, 200 + i)
        ow.cell(r, 3, 300 + i); nw.cell(r, 5, 300 + i)
        ow.cell(r, 4, 400 + i); nw.cell(r, 6, 400 + i)
        if i == 0:
            nw.cell(r, 4, 2000 + i)
            ow.cell(r, 3, 0.30000000000000004); nw.cell(r, 5, 0.3)
            ow.cell(r, 4, '=1+1'); nw.cell(r, 6, 2)
        elif i == 1:
            ow.cell(r, 2, '=1+1'); nw.cell(r, 4, '=2+2')
            ow.cell(r, 3, '=A1'); nw.cell(r, 5, '=a1')
            ow.cell(r, 4, '=SUM(A1:A1)'); nw.cell(r, 6, '=SUM(B1:B1)')
        elif i == 2:
            if HAS_RICH:
                ow.cell(r, 2, 'AB')
                nw.cell(r, 4, CellRichText(TextBlock(InlineFont(b=True), 'AB')))
                ow.cell(r, 3, CellRichText(TextBlock(InlineFont(rFont='微软雅黑'), 'AB')))
                nw.cell(r, 5, CellRichText([TextBlock(InlineFont(rFont='微软雅黑'), 'A'),
                                            TextBlock(InlineFont(rFont='微软雅黑'), 'B')]))
        elif i == 3:
            ow.cell(r, 2).font = Font(name='宋体')
            nw.cell(r, 4).font = Font(name='微软雅黑')
            ow.cell(r, 3).font = Font(name='微软雅黑')
            nw.cell(r, 5).font = Font(name='Microsoft YaHei')
        elif i == 4:
            ow.cell(r, 2).fill = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
            nw.cell(r, 4).fill = PatternFill(fill_type='solid', start_color='0000FF', end_color='0000FF')
            ow.cell(r, 3).fill = PatternFill(fill_type='solid', start_color=Color(theme=5), end_color=Color(theme=5))
            nw.cell(r, 5).fill = PatternFill(fill_type='solid', start_color='FF4472C4', end_color='FF4472C4')
            ow.cell(r, 4).fill = PatternFill(fill_type='solid', start_color=Color(theme=5, tint=-0.25), end_color=Color(theme=5, tint=-0.25))
            nw.cell(r, 6).fill = PatternFill(fill_type='solid', start_color='FF2F5597', end_color='FF2F5597')
        elif i == 5:
            thin = Side(style='thin', color='000000')
            nw.cell(r, 4).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        elif i == 6:
            nw.cell(r, 4).alignment = Alignment(horizontal='center')
        elif i == 7:
            ow.cell(r, 2).number_format = '0'
            nw.cell(r, 4).number_format = '0.00'
            ow.cell(r, 4, 0.25); nw.cell(r, 6, 0.25)
            ow.cell(r, 4).number_format = '0'
            nw.cell(r, 6).number_format = '0.00'
        elif i == 9:
            nw.row_dimensions[r].height = 30
        elif i == 10:
            nw.column_dimensions['B'].width = 30
            nw.column_dimensions['C'].width = 25
        elif i == 11:
            if HAS_PIL:
                buf = io.BytesIO()
                PILImage.new('RGB', (10, 10), (255, 0, 0)).save(buf, format='PNG')
                buf.seek(0)
                nw.add_image(XLImage(buf), 'C%d' % r)
        elif i == 12:
            if CellIsRule is not None:
                nw.conditional_formatting.add('D%d' % r, CellIsRule(operator='greaterThan', formula=['5'],
                                                                    fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')))
    return old, new


def _b1_shift_asserts():
    A = []
    for i in range(13):
        r = 2 + i
        if i == 8:
            A.append(('Sheet1', 'A%d' % r, 'DIFF', ['单元格删除']))
            A.append(('Sheet1', 'C%d' % r, 'DIFF', ['合并新增']))
        elif i == 9:
            A.append(('Sheet1', 'A%d' % r, 'DIFF', ['单元格删除']))
            A.append(('Sheet1', 'A%d' % r, 'PRESENT', ['行高变化']))
        elif i == 10:
            A.append(('Sheet1', 'C1', 'DIFF', ['列宽变化']))
            A.append(('Sheet1', 'B1', 'PRESENT', ['列宽变化']))
            A.append(('Sheet1', 'B%d' % r, 'DIFF', ['单元格删除']))
        elif i == 11:
            A.append(('Sheet1', 'A%d' % r, 'SAME', ['单元格删除']))
            if HAS_PIL:
                A.append(('Sheet1', 'C%d' % r, 'DIFF', ['图片新增', '图片变动', '图片尺寸变化']))
            else:
                A.append(('Sheet1', 'C%d' % r, 'NODIFF', None))
        elif i == 12:
            A.append(('Sheet1', 'A%d' % r, 'SAME', ['单元格删除']))
            A.append(('Sheet1', 'D%d' % r, 'DIFF', ['条件格式新增', '条件格式删除', '条件格式变化']))
        else:
            A.append(('Sheet1', 'A%d' % r, 'SAME', ['单元格删除']))
            A.append(('Sheet1', 'B%d' % r, 'DIFF', ['单元格删除']))
            A.append(('Sheet1', 'C%d' % r, 'SAME', ['内容变化']))
            if i in (0, 7):
                A.append(('Sheet1', 'D%d' % r, 'DIFF', ['内容变化']))
            else:
                A.append(('Sheet1', 'D%d' % r, 'SAME', ['内容变化']))
    return A


def _b1_cases():
    out = []
    layouts = [('普通', _b1_normal_build(), _b1_normal_asserts()),
               ('shift', _b1_shift_build(), _b1_shift_asserts())]
    for layout, (o, n), asserts in layouts:
        for exp in ('same', 'different'):
            if layout == '普通':
                if exp == 'same':
                    rules = [(_range_ds('Sheet1', 'L2ANCHOR', 1, 0, 13, 4), [(t, 'same') for t in CHECK_TYPES])]
                else:
                    rules = [(_range_ds('Sheet1', 'L2ANCHOR', 1 + i, 0, 1, 4), [(t, 'different')])
                             for i, t in enumerate(CHECK_TYPES)]
            else:
                rules = [(_shift_ds('Sheet1', 'L2ANCHOR', 2, str(2 + i)), [(t, exp)])
                         for i, t in enumerate(CHECK_TYPES)]
            out.append({'batch': 'B1', 'name': '%s规则·期望%s' % (layout, exp),
                        'build': (lambda o=o, n=n, rules=rules, asserts=asserts:
                                  {'old': o, 'new': n, 'rules': rules, 'asserts': asserts})})
    return out


# ------------------------------------------------------------
# B2 干扰矩阵
# ------------------------------------------------------------
def _b2_build(diff_types, rule_checks, expect, opts=None, make_diff=True):
    old = Workbook(); ow = old.active; ow.title = 'Sheet1'
    new = Workbook(); nw = new.active; nw.title = 'Sheet1'
    for ws in (ow, nw):
        ws['A1'] = 'L2ANCHOR'
        ws['A2'] = 10; ws['B2'] = 'keep'
    if make_diff:
        _fill_diff(ow, nw, 2, 1, diff_types)
    rules = [(_range_ds('Sheet1', 'L2ANCHOR', 1, 0, 2, 2), [(t, expect) for t in rule_checks])]
    return {'old': old, 'new': new, 'rules': rules, 'opts': opts or {},
            'asserts': [('Sheet1', 'A2', 'DIFF', None), ('Sheet1', 'B2', 'NODIFF', None)]}


def _b2_cases():
    from itertools import combinations
    out = []
    avail = [t for t in CHECK_TYPES if t in ('value', 'formula', 'rich_text', 'font', 'fill', 'border',
                                             'alignment', 'number_format', 'merged_cells', 'row_height',
                                             'col_width', 'images', 'conditional_format')]
    if not HAS_RICH:
        avail = [t for t in avail if t != 'rich_text']
    if not HAS_PIL:
        avail = [t for t in avail if t != 'images']
    for i, j in combinations(range(len(avail)), 2):
        t1, t2 = avail[i], avail[j]
        cs = [t1, t2]
        out.append({'batch': 'B2', 'name': '两两 %s+%s·same' % (t1, t2),
                    'build': (lambda cs=cs: _b2_build([cs[0]], cs, 'same'))})
    # 单格三差异（安全组合，value/formula 不同时出现）
    triples = [
        ('value', 'font', 'fill'), ('value', 'border', 'alignment'),
        ('value', 'number_format', 'conditional_format'), ('value', 'rich_text', 'images'),
        ('formula', 'font', 'fill'), ('formula', 'border', 'alignment'),
        ('formula', 'number_format', 'conditional_format'), ('formula', 'rich_text', 'images'),
        ('font', 'fill', 'border'), ('alignment', 'number_format', 'conditional_format'),
        ('rich_text', 'images', 'merged_cells'), ('value', 'merged_cells', 'conditional_format'),
        ('formula', 'merged_cells', 'row_height'), ('value', 'row_height', 'col_width'),
        ('font', 'alignment', 'conditional_format'), ('fill', 'number_format', 'rich_text'),
        ('formula', 'images', 'conditional_format'), ('value', 'font', 'conditional_format'),
    ]
    triples = [tp for tp in triples
               if all(t in avail for t in tp)]
    for tp in triples:
        out.append({'batch': 'B2', 'name': '单格多差 %s·same' % '+'.join(tp),
                    'build': (lambda tp=tp: _b2_build(list(tp), list(tp), 'same'))})
        out.append({'batch': 'B2', 'name': '单格多差 %s·different' % '+'.join(tp),
                    'build': (lambda tp=tp: _b2_build(list(tp), list(tp), 'different'))})
    return out


def _b2_full13_build():
    old = Workbook(); ow = old.active; ow.title = 'Sheet1'
    new = Workbook(); nw = new.active; nw.title = 'Sheet1'
    for ws in (ow, nw):
        ws['A1'] = 'L2ANCHOR'; ws['B1'] = 'keep'
    for i in range(13):
        r = 2 + i
        ow.cell(r, 1, 100 + i); nw.cell(r, 1, 100 + i)
        ow.cell(r, 2, 200 + i); nw.cell(r, 2, 200 + i)
        t = CHECK_TYPES[i]
        if i == 8:
            nw.merge_cells('B%d:B%d' % (r, r + 1))
            continue
        if i == 10:
            nw.column_dimensions['B'].width = 25
            continue
        _fill_diff(ow, nw, r, 1, [t])
    return old, new


def _b2_full13_asserts(kind='DIFF'):
    A = []
    for i in range(13):
        r = 2 + i
        if i == 8:
            if kind == 'NODIFF':
                A.append(('Sheet1', 'A%d' % r, 'NODIFF', None))
            continue
        A.append(('Sheet1', 'A%d' % r, kind, None))
    if kind == 'DIFF':
        A.append(('Sheet1', 'B10', 'DIFF', ['合并新增']))
        A.append(('Sheet1', 'B1', 'DIFF', ['列宽变化']))
    else:
        A.append(('Sheet1', 'B10', 'NODIFF', None))
        A.append(('Sheet1', 'B1', 'NODIFF', None))
    return A


def _b2_full_cases():
    out = []
    o, n = _b2_full13_build()
    rules = [(_range_ds('Sheet1', 'L2ANCHOR', 1, 0, 13, 2), [(t, 'same') for t in CHECK_TYPES])]
    out.append({'batch': 'B2', 'name': '全开13·每格单项差·same',
                'build': (lambda o=o, n=n, rules=rules:
                          {'old': o, 'new': n, 'rules': rules, 'asserts': _b2_full13_asserts('DIFF')})})
    return out


# ------------------------------------------------------------
# B3 全局交叉
# ------------------------------------------------------------
def _b3_cases():
    out = []
    ds = _range_ds('Sheet1', 'L2ANCHOR', 1, 0, 2, 2)
    # 1-6: value 专项（有差）
    combos = [
        ('B3-1 全局开·规则value same', [('value', 'same')], {}, 'A2', 'DIFF'),
        ('B3-2 全局开·规则value different', [('value', 'different')], {}, 'A2', 'DIFF'),
        ('B3-3 全局关·规则value same', [('value', 'same')], {'value': False}, 'A2', 'NODIFF'),
        ('B3-4 全局关·规则value different', [('value', 'different')], {'value': False}, 'A2', 'NODIFF'),
        ('B3-5 全局开·规则仅font same', [('font', 'same')], {}, 'A2', 'SAME'),
        ('B3-6 全局开·空检查规则', [], {}, 'A2', 'SAME'),
    ]
    for name, checks, opts, addr, kind in combos:
        exp = checks[0][1] if checks else 'same'
        rule_types = [t for t, e in checks]
        out.append({'batch': 'B3', 'name': name,
                    'build': (lambda rule_types=rule_types, exp=exp, opts=opts, addr=addr, kind=kind:
                              dict(_b2_build(['value'], rule_types, exp, opts),
                                   **{'asserts': [('Sheet1', addr, kind, None)]}))})
    # 7-8: 无差文件
    for name, checks in (('B3-7 无差+规则same', [('value', 'same')]),
                         ('B3-8 无差+规则different', [('value', 'different')])):
        out.append({'batch': 'B3', 'name': name,
                    'build': (lambda checks=checks:
                              dict(_b2_build([], [t for t, e in checks], checks[0][1]),
                                   **{'asserts': [('Sheet1', 'A2', 'NODIFF', None)]}))})
    # B3-B: 全局 13 项全关 + 规则 13 全开 different
    o, n = _b2_full13_build()
    rules = [(_range_ds('Sheet1', 'L2ANCHOR', 1, 0, 13, 2), [(t, 'different') for t in CHECK_TYPES])]
    allopt = {t: False for t in CHECK_TYPES}
    out.append({'batch': 'B3', 'name': 'B3-B 全局关13项+规则13开different',
                'build': (lambda o=o, n=n, rules=rules, allopt=allopt:
                          {'old': o, 'new': n, 'rules': rules, 'opts': allopt,
                           'asserts': _b2_full13_asserts('NODIFF')})})
    # B3-C: 无规则
    out.append({'batch': 'B3', 'name': 'B3-C 无规则·值差',
                'build': (lambda: dict(_b2_build(['value'], ['value'], 'same'),
                                       **{'rules': None, 'asserts': [('Sheet1', 'A2', 'PRESENT', None)]}))})
    return out


def _b4_base():
    """B4/B5 共用基础文件对：A1=锚点文字, B1=5, 数据格 A2"""
    old = Workbook(); ow = old.active; ow.title = 'Sheet1'
    new = Workbook(); nw = new.active; nw.title = 'Sheet1'
    for ws in (ow, nw):
        ws['A1'] = 'L2ANCHOR'
        ws['B1'] = 5
    return ow, nw
# ------------------------------------------------------------
# B4 COM 分歧 12 case
# ------------------------------------------------------------
def _b4_cases():
    out = []

    def _mk(prog_diff, fields, kind='observe', label='', old_addr='A2', new_addr='A2'):
        return [{'kind': kind, 'label': label or old_addr, 'old_sheet': 'Sheet1', 'old_addr': old_addr,
                 'new_sheet': 'Sheet1', 'new_addr': new_addr, 'fields': fields, 'prog_diff': prog_diff}]

    def c(name, make, asserts, com):
        def build():
            ow, nw = _b4_base()
            make(ow, nw)
            return {'old': ow, 'new': nw, 'asserts': asserts, 'com': com}
        out.append({'batch': 'B4', 'name': name, 'build': build})

    # 1 公式缓存值（openpyxl 无缓存 → 值读取为 None；Excel 显示计算值）
    def m1(ow, nw):
        ow['A2'] = '=1+1'; nw['A2'] = '=2'
    c('B4-1 公式缓存值', m1,
      [('Sheet1', 'A2', 'PRESENT', None)],
      _mk(True, ['text'], 'observe', '公式显示值'))

    # 2 浮点尾差
    def m2(ow, nw):
        ow['A2'] = 0.30000000000000004; nw['A2'] = 0.3
    c('B4-2 浮点尾差', m2,
      [('Sheet1', 'A2', 'NODIFF', None)],
      _mk(False, ['text'], 'observe', '尾差显示'))

    # 3 数字格式（存储同值、格式不同 → 显示 '0' vs '0.25'）
    def m3(ow, nw):
        ow['A2'] = 0.25; nw['A2'] = 0.25
        ow['A2'].number_format = '0'; nw['A2'].number_format = '0.00'
    c('B4-3 数字格式显示', m3,
      [('Sheet1', 'A2', 'PRESENT', ['数字格式变化'])],
      _mk(True, ['text'], 'observe', '格式显示'))

    # 4 字体名中文 vs 英文（font_equiv 等价表）
    def m4(ow, nw):
        ow['A2'] = 'x'; nw['A2'] = 'x'
        ow['A2'].font = Font(name='微软雅黑')
        nw['A2'].font = Font(name='Microsoft YaHei')
    c('B4-4 字体别名等价', m4,
      [('Sheet1', 'A2', 'NODIFF', None)],
      _mk(False, ['font_name'], 'observe', '字体别名'))

    # 5 主题色 tint（程序回算 vs Excel 真值 ±1 色阶）
    def m5(ow, nw):
        ow['A2'] = 10; nw['A2'] = 10
        ow['A2'].fill = PatternFill(fill_type='solid', start_color=Color(theme=5, tint=-0.25), end_color=Color(theme=5, tint=-0.25))
        nw['A2'].fill = PatternFill(fill_type='solid', start_color='FF2F5597', end_color='FF2F5597')
    c('B4-5 主题tint±1', m5,
      [('Sheet1', 'A2', 'NODIFF', None)],
      _mk(False, ['fill_bgr'], 'observe', 'tint色'))

    # 6 indexed 调色板（indexed=23 即 808080）
    def m6(ow, nw):
        ow['A2'] = 10; nw['A2'] = 10
        ow['A2'].fill = PatternFill(fill_type='solid', start_color=Color(indexed=23), end_color=Color(indexed=23))
        nw['A2'].fill = PatternFill(fill_type='solid', start_color='FF808080', end_color='FF808080')
    c('B4-6 indexed等价', m6,
      [('Sheet1', 'A2', 'NODIFF', None)],
      _mk(False, ['fill_bgr'], 'observe', 'indexed色'))

    # 7 行高 默认 vs 显式15（显示相同 → FP 候选）
    def m7(ow, nw):
        nw.row_dimensions[2].height = 15
    c('B4-7 行高默认vs15', m7,
      [('Sheet1', 'A2', 'PRESENT', ['行高变化'])],
      _mk(True, ['row_height'], 'fp', '行高15'))

    # 8 列宽 默认 vs 显式8.43（显示相同 → FP 候选）
    def m8(ow, nw):
        nw.column_dimensions['B'].width = 8.43
    c('B4-8 列宽默认vs8.43', m8,
      [('Sheet1', 'B1', 'PRESENT', ['列宽变化'])],
      _mk(True, ['col_width'], 'fp', '列宽8.43', 'B1', 'B1'))

    # 9 富文本（样式差异 → 程序报差，COM 文本相同）
    def m9(ow, nw):
        ow['A2'] = 'AB'
        nw['A2'] = CellRichText(TextBlock(InlineFont(b=True), 'AB'))
    c('B4-9 富文本样式', m9,
      [('Sheet1', 'A2', 'PRESENT', None)],
      _mk(True, ['text'], 'observe', '富文本'))

    # 10 条件格式（程序报 CF 变化；COM 读不到 CF 效果 → 观察）
    def m10(ow, nw):
        ow['A2'] = 10; nw['A2'] = 10
        if CellIsRule is not None:
            nw.conditional_formatting.add('A2', CellIsRule(operator='greaterThan', formula=['5'],
                                                           fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'))
    c('B4-10 条件格式', m10,
      [('Sheet1', 'A2', 'PRESENT', ['条件格式新增', '条件格式删除', '条件格式变化'])],
      _mk(True, ['text'], 'observe', '条件格式'))

    # 11 合并单元格（旧合并、新不合并）
    def m11(ow, nw):
        ow['A2'] = 10; nw['A2'] = 10
        ow['B2'] = 10; nw['B2'] = 10
        ow.merge_cells('A2:B2')
    c('B4-11 合并删除', m11,
      [('Sheet1', 'A2', 'PRESENT', ['合并删除'])],
      _mk(True, ['text'], 'observe', '合并'))

    # 12 公式大小写（归一化 → 无差；显示相同）
    def m12(ow, nw):
        ow['A2'] = '=B1'; nw['A2'] = '=b1'
    c('B4-12 公式大小写', m12,
      [('Sheet1', 'A2', 'NODIFF', None)],
      _mk(False, ['text'], 'observe', '公式大小写'))

    return out
# ------------------------------------------------------------
# B5 主题/字体等价专项
# ------------------------------------------------------------
def _mk_b5(prog_diff, fields, kind='observe', label='', old_addr='A2', new_addr='A2'):
    return [{'kind': kind, 'label': label or old_addr, 'old_sheet': 'Sheet1', 'old_addr': old_addr,
             'new_sheet': 'Sheet1', 'new_addr': new_addr, 'fields': fields, 'prog_diff': prog_diff}]


def _b5_cases():
    out = []

    def c(name, make, asserts, com, need_reload=False):
        def build():
            ow, nw = _b4_base()
            make(ow, nw)
            if need_reload:
                fp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
                sp = fp.name; fp.close()
                nw.save(sp)
                _set_theme_accent1(sp, 'FF0000')
                nw = load_workbook(sp)
                os.unlink(sp)
            return {'old': ow, 'new': nw, 'asserts': asserts, 'com': com}
        out.append({'batch': 'B5', 'name': name, 'build': build})

    # 1 主题色 theme=5 vs 直接 RGB 4472C4（显示相同）
    def m1(ow, nw):
        ow['A2'] = 10; nw['A2'] = 10
        ow['A2'].fill = PatternFill(fill_type='solid', start_color=Color(theme=5), end_color=Color(theme=5))
        nw['A2'].fill = PatternFill(fill_type='solid', start_color='FF4472C4', end_color='FF4472C4')
    c('B5-1 主题色vsRGB', m1,
      [('Sheet1', 'A2', 'NODIFF', None)],
      _mk_b5(False, ['fill_bgr'], 'observe', '主题色'))

    # 2 同 theme 索引、主题表 accent1 不同（旧4472C4/新FF0000）
    def m2(ow, nw):
        ow['A2'] = 10; nw['A2'] = 10
        ow['A2'].fill = PatternFill(fill_type='solid', start_color=Color(theme=5), end_color=Color(theme=5))
        nw['A2'].fill = PatternFill(fill_type='solid', start_color=Color(theme=5), end_color=Color(theme=5))
    c('B5-2 主题表不同', m2,
      [('Sheet1', 'A2', 'PRESENT', ['填充变化'])],
      _mk_b5(True, ['fill_bgr'], 'observe', '跨主题表'), need_reload=True)

    # 3 tint -0.25 vs Excel 真值 2F5597（±1 容差）
    def m3(ow, nw):
        ow['A2'] = 10; nw['A2'] = 10
        ow['A2'].fill = PatternFill(fill_type='solid', start_color=Color(theme=5, tint=-0.25), end_color=Color(theme=5, tint=-0.25))
        nw['A2'].fill = PatternFill(fill_type='solid', start_color='FF2F5597', end_color='FF2F5597')
    c('B5-3 tint±1', m3,
      [('Sheet1', 'A2', 'NODIFF', None)],
      _mk_b5(False, ['fill_bgr'], 'observe', 'tint色'))

    # 4 indexed=23 vs RGB 808080
    def m4(ow, nw):
        ow['A2'] = 10; nw['A2'] = 10
        ow['A2'].fill = PatternFill(fill_type='solid', start_color=Color(indexed=23), end_color=Color(indexed=23))
        nw['A2'].fill = PatternFill(fill_type='solid', start_color='FF808080', end_color='FF808080')
    c('B5-4 indexed等价', m4,
      [('Sheet1', 'A2', 'NODIFF', None)],
      _mk_b5(False, ['fill_bgr'], 'observe', 'indexed色'))

    # 5 字体 中文名 vs 英文名（font_equiv）
    def m5(ow, nw):
        ow['A2'] = 'x'; nw['A2'] = 'x'
        ow['A2'].font = Font(name='微软雅黑')
        nw['A2'].font = Font(name='Microsoft YaHei')
    c('B5-5 字体别名', m5,
      [('Sheet1', 'A2', 'NODIFF', None)],
      _mk_b5(False, ['font_name'], 'observe', '字体别名'))

    # 6 主题字体 scheme=minor vs 直接名 Calibri（显示同为 Calibri）
    def m6(ow, nw):
        ow['A2'] = 'x'; nw['A2'] = 'x'
        ow['A2'].font = Font(scheme='minor')
        nw['A2'].font = Font(name='Calibri')
    c('B5-6 主题字体scheme', m6,
      [('Sheet1', 'A2', 'NODIFF', None)],
      _mk_b5(False, ['font_name'], 'observe', 'scheme字体'))

    # 7 组合等价：Calibri+主题色 vs scheme+theme5
    def m7(ow, nw):
        ow['A2'] = 10; nw['A2'] = 10
        ow['A2'].font = Font(name='Calibri', size=11)
        ow['A2'].fill = PatternFill(fill_type='solid', start_color='FF4472C4', end_color='FF4472C4')
        nw['A2'].font = Font(scheme='minor', size=11)
        nw['A2'].fill = PatternFill(fill_type='solid', start_color=Color(theme=5), end_color=Color(theme=5))
    c('B5-7 组合等价', m7,
      [('Sheet1', 'A2', 'NODIFF', None)],
      _mk_b5(False, ['font_name', 'fill_bgr'], 'observe', '组合等价'))

    # 8 真差异：微软雅黑 vs 宋体（防过度豁免）
    def m8(ow, nw):
        ow['A2'] = 'x'; nw['A2'] = 'x'
        ow['A2'].font = Font(name='微软雅黑')
        nw['A2'].font = Font(name='宋体')
    c('B5-8 真字体差异', m8,
      [('Sheet1', 'A2', 'PRESENT', ['字体变化'])],
      _mk_b5(True, ['font_name'], 'observe', '真差异'))

    return out


# ------------------------------------------------------------
# 汇总
# ------------------------------------------------------------
def suite():
    out = []
    out.extend(_b1_cases())
    out.extend(_b2_cases())
    out.extend(_b2_full_cases())
    out.extend(_b3_cases())
    out.extend(_b4_cases())
    out.extend(_b5_cases())
    return out

