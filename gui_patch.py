# -*- coding: utf-8 -*-
"""
规则编辑窗「数据源测试」补丁 v2.2（引擎同源版）：
- 测试按钮不再自写范围计算，全部复用 main.py 引擎逻辑：
  * offset/intersection/range：DataLocator.locate_all()（与 _apply_rule_filter 相同调用），旧/新双侧定位
  * shift：OpenpyxlComparer.shift_scope()（引擎运行时同款提取函数，v3.99）
- 点击测试 → 直接执行：引擎计算 → 文件未开则询问并打开（40s等待，界面不卡）
  → 弹出主界面同款 ComCheckDialog 确认窗 → 确认后跳转（旧→新，只选数据区）
- 跳转复用主程序 jump_to_excel（支持多区域联合选中）；成功无弹窗，失败才提示
不修改 main.py 行为，仅运行时替换类方法。
"""
import os, re, time
import openpyxl
import main
from tkinter import messagebox

_VIEWER = None   # 主窗口（DiffViewer）全局登记，保证测试按钮能找到 jump_to_excel


def _find_viewer(widget):
    """全局登记优先；兜底向上遍历 parent/master"""
    global _VIEWER
    if _VIEWER is not None:
        return _VIEWER
    cur = widget
    while cur is not None:
        if hasattr(cur, 'jump_to_excel'):
            return cur
        cur = getattr(cur, 'parent', None) or getattr(cur, 'master', None)
    return None


def _group_addresses(addrs):
    """把单元格地址列表合并为 Excel 联合选择串，如 'B48:C49,E5'。返回 (联合串, 格数)"""
    pts = []
    for a in addrs:
        m = re.match(r'^([A-Za-z]+)(\d+)$', str(a))
        if not m:
            continue
        c = main.column_index_from_string(m.group(1).upper())
        r = int(m.group(2))
        pts.append((r, c))
    if not pts:
        return '', 0
    pts.sort()
    rows = {}
    for r, c in pts:
        rows.setdefault(r, []).append(c)
    segs = []
    for r in sorted(rows):
        cols = sorted(set(rows[r]))
        start = prev = cols[0]
        for c in cols[1:]:
            if c == prev + 1:
                prev = c
            else:
                segs.append((r, start, prev))
                start = prev = c
        segs.append((r, start, prev))
    segs.sort()
    ranges = []
    i = 0
    n = len(segs)
    while i < n:
        r, c1, c2 = segs[i]
        j = i
        while (j + 1 < n and segs[j+1][1] == c1 and segs[j+1][2] == c2
               and segs[j+1][0] == segs[j][0] + 1):
            j += 1
        top = segs[i][0]
        bottom = segs[j][0]
        a1 = main.cell_address(c1, top)
        a2 = main.cell_address(c2, bottom)
        ranges.append(a1 if a1 == a2 else '%s:%s' % (a1, a2))
        i = j + 1
    return ','.join(ranges), len(pts)


def _bbox_of(addrs):
    """整体包围框，如 'B2:F9'（超量兜底用）"""
    pts = []
    for a in addrs:
        m = re.match(r'^([A-Za-z]+)(\d+)$', str(a))
        if m:
            pts.append((int(m.group(2)), main.column_index_from_string(m.group(1).upper())))
    if not pts:
        return ''
    rows = [p[0] for p in pts]; cols = [p[1] for p in pts]
    return '%s:%s' % (main.cell_address(min(cols), min(rows)), main.cell_address(max(cols), max(rows)))


def _parse_exclude(self, text):
    """增强版排除解析：支持 单格A5 / 相对[1,0] / 区间A1-A9、B2-D2、A1:A9（自动展开）"""
    text = (text or '').replace('，', ',')
    if not text.strip():
        return []
    parts = [p.strip() for p in text.split(',') if p.strip()]
    exclude = []
    for p in parts:
        if re.match(r'^[A-Za-z]+\d+$', p):
            exclude.append(p.upper())
        elif re.match(r'^\[\d+,\d+\]$', p):
            inner = p[1:-1].split(',')
            exclude.append([int(inner[0]), int(inner[1])])
        elif re.match(r'^[A-Za-z]+\d+[-:][A-Za-z]+\d+$', p):
            m = re.match(r'^([A-Za-z]+)(\d+)[-:]([A-Za-z]+)(\d+)$', p)
            if m:
                c1_s, r1_s, c2_s, r2_s = m.groups()
                c1 = main.column_index_from_string(c1_s.upper())
                c2 = main.column_index_from_string(c2_s.upper())
                r1, r2 = int(r1_s), int(r2_s)
                cnt = 0
                if c1 == c2:
                    col = c1_s.upper()
                    for r in range(min(r1, r2), max(r1, r2) + 1):
                        exclude.append('%s%d' % (col, r)); cnt += 1
                        if cnt > 5000:
                            break
                elif r1 == r2:
                    row = r1
                    for c in range(min(c1, c2), max(c1, c2) + 1):
                        exclude.append('%s%d' % (main.get_column_letter(c), row)); cnt += 1
                        if cnt > 5000:
                            break
                else:
                    for r in range(min(r1, r2), max(r1, r2) + 1):
                        for c in range(min(c1, c2), max(c1, c2) + 1):
                            exclude.append('%s%d' % (main.get_column_letter(c), r)); cnt += 1
                            if cnt > 5000:
                                break
    return exclude


def _ds_assemble(self):
    """与 on_ok 完全一致的 data_source 组装（去掉检查项部分）"""
    si = self.search_in_var.get().strip().upper().replace('$', '')
    try:
        ds_si = si if main.DataLocator._parse_area(si) else 'all'
    except Exception:
        ds_si = 'all'
    ds = {'name': self.rule_name_var.get(), 'sheet': self.sheet_var.get(),
          'anchor': {'text': self.anchor_text_var.get()}, 'search_in': ds_si,
          'mode': self.mode_var.get()}
    mode = self.mode_var.get()
    if mode == 'offset':
        ds['target'] = {'row_offset': self._to_int(self.offset_row_var), 'col_offset': self._to_int(self.offset_col_var)}
    elif mode == 'intersection':
        ds['row_anchor'] = {'text': self.row_anchor_text_var.get().strip(), 'search_in': ds_si}
        ds['col_anchor'] = {'text': self.col_anchor_text_var.get().strip(), 'search_in': ds_si}
    elif mode == 'range':
        ds['target'] = {'row_offset': self._to_int(self.range_row_offset_var),
                        'col_offset': self._to_int(self.range_col_offset_var),
                        'row_count': self._to_int(self.range_row_count_var, 1) or 1,
                        'col_count': self._to_int(self.range_col_count_var, 1) or 1,
                        'exclude': _parse_exclude(self, self.range_exclude_var.get())}
    elif mode == 'shift':
        ds['header_target'] = {'row_offset': self._to_int(self.sh_ro_var),
                               'col_offset': self._to_int(self.sh_co_var),
                               'row_count': max(1, self._to_int(self.sh_rc_var, 1)),
                               'col_count': max(1, self._to_int(self.sh_cc_var, 20))}
        ds['rows'] = self.sh_rows_var.get().strip()
        ds['shift_offset'] = self._to_int(self.sh_offset_var)
    return ds


def _jump_once(self, file_path, sheet_name, addr):
    """复用主程序 jump_to_excel（引擎同款跳转，支持联合多区）"""
    viewer = _find_viewer(self)
    if viewer is None:
        return False, '无法定位主窗口（jump_to_excel 不可用）'
    try:
        ok, err = viewer.jump_to_excel(file_path, sheet_name, addr)
        if ok == 'opened':
            return True, None
        return False, err
    except Exception as e:
        return False, str(e)


def _ds_test(self):
    try:
        mode = self.mode_var.get() or 'offset'
        sheet = self.sheet_var.get()          # 保留原始名称（引擎按精确名称匹配）
        if not sheet:
            messagebox.showwarning('测试', '请先选择 Sheet'); return
        old_path = self.old_path; new_path = self.new_path
        if not old_path or not os.path.isfile(old_path):
            messagebox.showwarning('测试', '参考报告（旧）文件不存在，请先选择'); return
        if not new_path or not os.path.isfile(new_path):
            messagebox.showwarning('测试', '待检报告（新）文件不存在，请先选择'); return
        ds = _ds_assemble(self)
        self.config(cursor='watch')
        try:
            owb = openpyxl.load_workbook(old_path, data_only=False)
            wb = openpyxl.load_workbook(new_path, data_only=False)
            if sheet not in owb.sheetnames:
                messagebox.showwarning('测试', '引擎将跳过：旧文件无该 Sheet（实际: %s）' % '、'.join(owb.sheetnames[:15])); return
            if sheet not in wb.sheetnames:
                messagebox.showwarning('测试', '引擎将跳过：新文件无该 Sheet（实际: %s）' % '、'.join(wb.sheetnames[:15])); return
            if mode == 'shift':
                if ds.get('old_sheet') or ds.get('old_anchor'):
                    messagebox.showwarning('测试', '旧版双区域shift结构已不支持，请重新编辑该规则'); return
                tmp_rule = main.CheckRule(rule_name=ds.get('name', '测试'), data_source=ds)
                logs = []
                eng = main.OpenpyxlComparer(old_path, new_path)
                scope = eng.shift_scope(owb, wb, tmp_rule, log_fn=lambda m: logs.append(m))
                if not scope:
                    reason = '\n'.join(logs) if logs else '（无日志：Sheet 必须同名并存于新旧文件，或规则为旧版结构）'
                    messagebox.showwarning('测试', '引擎跳过该规则（与运行时一致）：\n' + reason); return
                pairs = scope['pairs']; rowset = scope['rowset']
                old_cells = [main.cell_address(oc, r) for oc, nc in pairs for r in sorted(rowset)]
                new_cells = [main.cell_address(nc, r) for oc, nc in pairs for r in sorted(rowset)]
                old_union, old_cnt = _group_addresses(old_cells)
                new_union, new_cnt = _group_addresses(new_cells)
                old_note = new_note = ''
                if old_cnt > 2000 or len(old_union.split(',')) > 300:
                    old_union = _bbox_of(old_cells); old_cnt = len(old_cells); old_note = '（量大，整体选中）'
                if new_cnt > 2000 or len(new_union.split(',')) > 300:
                    new_union = _bbox_of(new_cells); new_cnt = len(new_cells); new_note = '（量大，整体选中）'
                jump_addrs = [(old_path, sheet, old_union), (new_path, sheet, new_union)]
            else:
                loc = main.DataLocator(); loc.rules = [ds]
                name = ds.get('name', '')
                old_res = loc.locate_all(owb).get(name)
                new_res = loc.locate_all(wb).get(name)
                if not isinstance(old_res, dict) or not isinstance(new_res, dict):
                    messagebox.showwarning('测试', '引擎将跳过：定位结果缺失'); return
                for res, label in ((old_res, '旧'), (new_res, '新')):
                    if 'error' in res:
                        messagebox.showwarning('测试', '引擎将跳过：%s文件定位失败: %s' % (label, res['error'])); return
                def _addrs_of(res):
                    return res.get('addresses') or ([res.get('address')] if res.get('address') else [])
                old_addrs = [a for a in _addrs_of(old_res) if a]
                new_addrs = [a for a in _addrs_of(new_res) if a]
                if not old_addrs:
                    messagebox.showwarning('测试', '引擎将跳过：旧文件定位结果为空'); return
                old_union, old_cnt = _group_addresses(old_addrs)
                new_union, new_cnt = _group_addresses(new_addrs)
                old_note = new_note = ''
                if old_cnt > 2000 or len(old_union.split(',')) > 300:
                    old_union = _bbox_of(old_addrs); old_cnt = len(old_addrs); old_note = '（量大，整体选中）'
                if new_cnt > 2000 or len(new_union.split(',')) > 300:
                    new_union = _bbox_of(new_addrs); new_cnt = len(new_addrs); new_note = '（量大，整体选中）'
                jump_addrs = [(old_path, sheet, old_union), (new_path, sheet, new_union)]
            # ---- 文件开启检查 + COM 通道确认（照搬主界面 _prep_advanced_audit 流程）----
            viewer = _find_viewer(self)
            if viewer is None:
                messagebox.showwarning('测试', '无法定位主窗口（jump_to_excel 不可用）'); return
            miss = []
            paths = viewer._excel_open_paths()
            if paths is None:
                miss = [old_path, new_path]
            else:
                if main.normalize_path(old_path) not in paths: miss.append(old_path)
                if main.normalize_path(new_path) not in paths: miss.append(new_path)
            # 文件锁兜底：COM 没找到（可能返回了其他 Excel 实例），但文件已被锁定说明确实打开了
            if miss and viewer._is_file_locked(old_path) and viewer._is_file_locked(new_path):
                miss = []
            if miss:
                choice = messagebox.askyesnocancel('测试', '跳转需通过 Excel COM 调用显示层，报告文件尚未打开。\n是＝立即打开报告并继续；否/取消＝中止本次跳转。')
                if choice is None or not choice:
                    return
                for fp in miss:
                    try:
                        os.startfile(fp)
                    except Exception as e:
                        messagebox.showerror('错误', '无法打开文件：%s' % e); return
                ok_all = False; deadline = time.time() + 40
                while time.time() < deadline:
                    time.sleep(1)
                    try:
                        self.update()
                    except Exception:
                        pass
                    ps = viewer._excel_open_paths()
                    if ps is not None and main.normalize_path(old_path) in ps and main.normalize_path(new_path) in ps:
                        ok_all = True; break
                    if viewer._is_file_locked(old_path) and viewer._is_file_locked(new_path):
                        ok_all = True; break
                if not ok_all and not messagebox.askyesno('测试', '报告打开超时，是否仍继续跳转？'):
                    return
            # COM 通道确认窗（主界面同款：检查COM通道 → 确认开启）
            dlg = main.ComCheckDialog(viewer.root)
            if dlg.result is not True:
                return
            # ---- 跳转：先旧后新，成功不弹窗 ----
            errs = []
            for fp, sh, ad in jump_addrs:
                ok, err = _jump_once(self, fp, sh, ad)
                if not ok:
                    errs.append('%s: %s' % (main.normalize_path(fp), err))
            if errs:
                messagebox.showwarning('跳转失败', '\n'.join(errs))
        finally:
            self.config(cursor='')
    except Exception as e:
        messagebox.showerror('测试异常', str(e))


def apply():
    global _VIEWER
    # 登记主窗口：DiffViewer 创建时记录实例，测试按钮由此拿到 jump_to_excel
    _orig_dv_init = main.DiffViewer.__init__

    def _patched_dv_init(self, *args, **kwargs):
        _orig_dv_init(self, *args, **kwargs)
        global _VIEWER
        _VIEWER = self
    main.DiffViewer.__init__ = _patched_dv_init

    _orig_build_ui = main.RuleEditorDialog._build_ui

    def _patched_build_ui(self):
        _orig_build_ui(self)
        try:
            btn_row = main.tb.Frame(self.param_frame.master)
            btn_row.pack(fill='x', pady=(6, 2))
            main.tb.Button(btn_row, text="测试", bootstyle="success-outline", width=8,
                           command=self._ds_test).pack(side='left')
        except Exception:
            pass
    main.RuleEditorDialog._build_ui = _patched_build_ui
    main.RuleEditorDialog._ds_test = _ds_test
    main.RuleEditorDialog._parse_exclude = _parse_exclude   # 增强排除解析，同步影响真实规则保存
